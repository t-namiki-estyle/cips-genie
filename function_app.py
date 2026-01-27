import azure.functions as func
import azure.durable_functions as d_func
import logging
import urllib.request
import urllib.parse
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import base64
import io
import sys
import random
import traceback
from openpyxl import Workbook

import ast
import csv
import asyncio

# OSS
import requests
import pandas as pd
import markdown
from pydantic import BaseModel, Field
import openai

# 自作
from config import (
    LLM_REGISTRY,
    NON_CHAT_REGISTRY,
    COSMOS_CLIENT,
    VARIABLE_LIST,
    ENVIRONMENT_SELECTED,
)
from util import *
from utils.token import decode_id_token
from utils.user_auth.user_auth_manager import (
    UserDivisionFetchService,
    MenuPermissionService,
)

from i_style.aiohttp import AsyncHttpClient, http_post
from i_style.token import EntraIDTokenManager
from i_style.llm import AzureOpenAI, GeminiGenerate, ClaudeGenerate
from openai import AsyncAzureOpenAI

from prompt import (
    func_list,
    replace_dict,
    language_dict,
    ocr_prompt_list,
    ocr_prompt_list_business_pattern,
    ocr_prompt_list_shipping,
    query_prompt,
    process_prompt,
    choice_prompt,
    formatted_date,
    FavoritePromptManager,
    QueryList,
    LinkList,
    BASE_SYSTEM_CONTENT,
    WEB_SYSTEM_CONTENT,
    MAIL_SYSTEM_CONTENT,
    TEAMS_SYSTEM_CONTENT,
    CHAT_SYSTEM_CONTENT,
    CSE_SYSTEM_CONTENT,
    CSE_RESULT_SYSTEM_CONTENT,
    query_suffix,
    MINUTES_SYSTEM_CONTENT,
    whisper_options,
    transcribe_prompt_dict,
    TRANSLATION_SYSTEM_CONTENT,
    CRM_SYSTEM_CONTENT,
    CRM_USER_CONTENT,
    OCR_CSV_SYSTEM_CONTENT,
    SEARCH_GROUNDING_PROMPT,
    GOOGLE_CONTENT_SUMMARIZE_SYSTEM_CONTENT,
    QUERY_GENERATION_SYSTEM_CONTENT,
    LINK_SELECTION_SYSTEM_CONTENT,
)

from config import (
    GPT_API_VERSION,
    GPT4O_TRANSCRIBE_API_ENDPOINT,
    GPT4O_TRANSCRIBE_API_KEY,
    GPT4O_TRANSCRIBE_DEPLOYMENT_NAME,
    MCP_AGENT_URL,
    MCP_AGENT_API_KEY,
    GEMINI_DEFAULT_LABELS,
)
from replace_list import replace_list

from crm import csv_search
from whisper_bp import bp as whisper_bp
from zoom_bp import bp as zoom_bp
from audio_upload_bp import bp as audio_upload_bp  # これは元のままでOK
from box_bp import bp as box_bp
from log_bp import bp as log_bp, access_merchant_rate
from history_bp import bp as history_bp
from ocr_bp import bp as ocr_bp
from file_diff_bp import bp as file_diff_bp

from blueprints import blob_dl_bp, hanabi_bp, word_bp

from utils.enq_apis.api_call import EnqAPICall
from utils.enq_apis.authority_verification import AuthorityVerification

#
# global
#
MERCHANT_RATE = {"daily_user": 0, "next_update": "0001-01-01T00:00:00"}

GEMINI_IGNORE_SIGN = "<!-- gemini検索結果です。使用しません。 -->"

GPT5_MODEL_SERIES = {"gpt5-low": "low", "gpt5-medium": "medium", "gpt5-high": "high"}


########
# main #
########
# app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)
app = d_func.DFApp(http_auth_level=func.AuthLevel.FUNCTION)

# blueprints
# app.register_blueprint(chat_bp)
app.register_blueprint(whisper_bp)
app.register_blueprint(zoom_bp)
app.register_blueprint(audio_upload_bp)
app.register_blueprint(box_bp)
app.register_blueprint(log_bp)
app.register_blueprint(history_bp)
app.register_blueprint(ocr_bp)
app.register_blueprint(blob_dl_bp)
app.register_blueprint(file_diff_bp)
app.register_blueprint(hanabi_bp)
app.register_blueprint(word_bp)


# Genie
@app.route(route="genie", methods=("POST",))
async def genie(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Genie processed a request.")
    ########
    # init #
    ########
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    # apiの呼び出し
    history_base_url = os.environ.get("HISTORY_API_URL")
    history_api_key = os.environ.get("HISTORY_API_KEY")

    title = "No Title"

    # get client ip
    client_ip = req.headers.get("x-forwarded-for")

    # get json input
    req_json = req.get_json()
    mode = req_json["mode"]

    session_id = req_json.get("session_id")
    logging.info(f"session_id: {session_id}")

    # DBにアクセスする場合はid_tokenでの認証を実施する
    if session_id != None or mode == "box":
        try:
            id_token = req.params.get("id_token")
            keys = await EntraIDTokenManager.get_entra_openid_keys(async_http_client)
            upn, mail, _ = decode_id_token(id_token, keys)
        except Exception as e:
            logging.warning(f"token error: {e}")
            response = error_response(
                "認証に失敗しました。ページの再読み込みをお試しください。"
            )
            response["blobs"] = []

            return func.HttpResponse(json.dumps(response), status_code=200)  # 503
    else:
        upn = req_json.get("upn")
        mail = req_json.get("mail")

    # web, teams, mail, agentのどこから来たかログを取る
    try:
        sendFrom = req_json["from"]
        assert sendFrom in ["web", "teams", "mail", "agent"]
    except Exception as e:
        req_json["from"] = sendFrom = "web"

    try:
        model = req_json["model"]
        if model in GPT5_MODEL_SERIES:
            assert "gpt5" in LLM_REGISTRY.list_models()
        else:
            assert model in LLM_REGISTRY.list_models()
    except AssertionError as e:
        logging.warning("invalid model name")
        model = "gpt4.1"
    except:
        logging.warning("no model name")
        model = "gpt4.1"
    req_json["model"] = model

    logging.info(f"client_ip: {client_ip}")
    logging.info("mode: " + mode)
    logging.info(f"from: {sendFrom}")
    logging.info("upn: " + upn)
    logging.info(f"mail: {mail}")
    logging.info("model: " + model)

    try:
        messages = req_json["messages"]

        # 履歴追加用に待避
        new_message = deepcopy(messages[-1])
    except KeyError:
        messages = []
    except IndexError as e:
        logging.warning(f"IndexError: {e}")
        new_message = {"role": "user", "content": ""}

    if session_id != None:
        params = {"sessionId": session_id}

        if mode in {"inside", "minutes"}:
            url = f"{history_base_url}/api/history/{mode}/{upn}"
        else:
            url = f"{history_base_url}/api/history/genie/{upn}"

        api_name = "get_history"
        try:
            history_json = await async_http_client.get(
                url=url, api_key=history_api_key, params=params, process_name=api_name
            )
        except Exception as e:
            logging.warning(f"履歴の取得: {e}")
        else:
            if mode == "minutes":
                # 全てのmessageを汎用履歴から取得する
                messages = []
                transcription = ""
                for message in history_json["messages"]:
                    # 文字起こしか議事録かの確認
                    if message.get("model") in {"whisper", "gpt-4o-transcribe"}:
                        _type = "transcribe"
                    else:
                        _type = "minutes"

                    # 翻訳を行っている場合は翻訳を除去する
                    if message["role"] == "assistant" and _type == "transcribe":
                        for item in message["content"]:
                            if item["type"] == "text":
                                transcription += (
                                    item["text"]
                                    .split("\n" + "*" * 20 + "\n")[0]
                                    .strip()
                                    + "\n"
                                )

                # 複数改行の調整
                transcription = re.sub(r"\n{3,}", "\n\n", transcription)

                # 一つのメッセージとして必要な情報のみ追加
                messages.append(
                    {
                        "role": "user",
                        "content": [{"type": "text", "text": transcription}],
                    }
                )

            else:
                # 必要な情報のみ追加
                messages_history = []
                for message in history_json["messages"]:
                    for item in message["content"]:
                        if item["type"] == "text":
                            # GEMINI_IGNORE_SIGNが含まれている場合、その部分以降を削除
                            ignore_index = item["text"].find(GEMINI_IGNORE_SIGN)
                            if ignore_index != -1:
                                item["text"] = item["text"][:ignore_index]

                    message_dict = {
                        "role": message["role"],
                        "content": message["content"],
                    }
                    messages_history.append(message_dict)

                # 最新のmessageのみ、入力から受け取る
                messages = messages_history + messages[-1:]

    # ファイルアップロード機能の処理の追加
    formatted_messages, blobs = await messages2textMessages(
        upn, messages, async_http_client
    )
    blob_names = [blob["name"] for blob in blobs]

    user_input = next(
        (
            content["text"]
            for content in formatted_messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )

    logging.info(f"User input: {user_input}")
    # suffixの除去
    messages = filter_messages(formatted_messages, escape_strings="<br><hr>")
    ########
    # main #
    ########
    try:
        func_name = await Brain(messages=messages, mode=mode, func_list=func_list)
    except Exception as e:
        logging.warning(f"Brain error: {e}")
        response = error_response(
            "サーバーからの応答がありません。時間をおいてお試しください。"
        )
        response["blobs"] = blob_names
        return func.HttpResponse(json.dumps(response), status_code=200)  # 503
    logging.info("Brain function: " + func_name)
    user_input = messages[-1]["content"]

    # 0,1入力時のmessageの変更
    if user_input in ("0", "０", "0:", "０：", "1", "１", "1:", "１："):
        messages[-1] = change_user_input(messages)
        user_input = messages[-1]["content"]
        logging.info("Changed User input:" + user_input)

    if user_input[:2] in ("0", "０", "0:", "０：", "1", "１", "1:", "１："):
        messages[-1]["content"] = user_input[2:]
        user_input = user_input[2:]
        logging.info("Changed User input:" + user_input)

    # error
    if user_input == "error":
        message = "質問を見つけられませんでした。もう一度質問を入力してください。"
        response = error_response(message)
        response["blobs"] = blob_names
        return func.HttpResponse(json.dumps(response), status_code=200)  # 503

    # enq
    if func_name == "LLM_DOCS":
        func_name = "LLM_ENQ"

    # API KEYとURLの取得
    FUNCTION_API_KEY = os.environ.get(f"{func_name}_API_KEY")
    FUNCTION_URL = os.environ.get(f"{func_name}_URL")

    # API呼び出し
    try:
        req_json["messages"] = messages
        api_response = await http_post(
            json_data=req_json,
            url=FUNCTION_URL,
            api_key=FUNCTION_API_KEY,
            process_name=func_name,
        )
        api_response_text = api_response["choices"][0]["message"]["content"]
        if api_response["object"] == "error":
            status_code = 200  # 503
        else:
            status_code = 200
        logging.info("Assistant response: " + api_response_text)

        # ```html で始まるコードブロックを除去
        api_response["choices"][0]["message"]["content"] = re.sub(
            r"```([a-zA-Z0-9]+)?\s*\n?(.*?)```",
            html_replacer,
            api_response["choices"][0]["message"]["content"],
            flags=re.DOTALL,
        )

        # blobsの追加
        api_response["blobs"] = blob_names
    except Exception as e:
        logging.critical(f"{e}")

        # blobsの追加
        api_response = error_response(
            "サーバーからの応答がありません。時間をおいてお試しください。"
        )
        api_response["blobs"] = blob_names

    # 履歴の登録
    if session_id != None:
        submode_mapping = {"LLM_CHAT": "chat", "LLM_GOOGLE": "google"}
        submode = submode_mapping.get(func_name, "")

        request_data = {"items": []}

        if mode != "minutes":

            # 汎用履歴への登録用に最新のmessageの中にfileがある場合はblobに置換
            blob_contents_info = [
                {"type": "blob", "name": blob["name"], "file_name": blob["file_name"]}
                for blob in blobs
            ]

            content = new_message["content"]
            if isinstance(content, list):
                other_contents_info = [
                    item.copy() for item in content if item.get("type") != "file"
                ]
            elif isinstance(content, str):
                other_contents_info = [{"type": "text", "text": content}]

            converted_content = blob_contents_info + other_contents_info

            request_data["items"].append(
                {
                    "upn": upn,
                    "content": converted_content,
                    "role": "user",
                    "submode": submode,
                    "model": model,
                    "from": sendFrom,
                    "sessionId": session_id,
                }
            )

        content = {
            "type": "text",
            "text": api_response["choices"][0]["message"]["content"],
        }

        request_data["items"].append(
            {
                "upn": upn,
                "content": [content],
                "role": "assistant",
                "submode": submode,
                "model": model,
                "from": sendFrom,
                "sessionId": session_id,
            }
        )

        # history_base_url: https://itc-history-functions.azurewebsites.net
        if mode not in {"inside", "minutes"}:
            mode = "genie"
        url = f"{history_base_url}/api/history/{mode}"

        api_name = "add_history"
        try:
            response = await async_http_client.post(
                url=url,
                api_key=history_api_key,
                json_data=request_data,
                process_name=api_name,
            )
        except Exception as e:
            logging.warning(f"履歴の追加: {e}")
        else:
            api_response["session_id"] = response["sessionId"]
            title = response.get("title", "タイトルの取得に失敗しました。")

        api_response["title"] = title
    return func.HttpResponse(json.dumps(api_response))


# LLM_CHAT
@app.route(route="llm/chat", methods=("POST",))
async def llm_chat(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("LLM_CHAT processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    sendFrom = req_json["from"]

    model_name = req_json.get("model", "gpt4.1")

    messages = req_json["messages"]
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)

    ui_system_content = ""
    if sendFrom == "web":
        ui_system_content = WEB_SYSTEM_CONTENT
    if sendFrom == "mail":
        ui_system_content = MAIL_SYSTEM_CONTENT
    if sendFrom == "teams":
        ui_system_content = TEAMS_SYSTEM_CONTENT

    # prompt
    chat_system_content = BASE_SYSTEM_CONTENT + ui_system_content + CHAT_SYSTEM_CONTENT
    chat_system_content = chat_system_content.format(formatted_date=formatted_date())

    ########
    # main #
    ########
    chat_messages = change_system_content(messages, chat_system_content)
    if sendFrom in [
        "teams",
    ]:
        for content in chat_messages[-1]["content"]:
            if content["type"] == "text":
                content["text"] += "\nHTML形式で回答してください。"
    try:
        if model_name.startswith("claude-"):
            # Claude モデルの場合
            chat_response = await ClaudeGenerate(
                chat_messages,
                model_name=model_name,
                max_retries=2,
                timeout=230,
                raise_for_error=False,
                registry=LLM_REGISTRY,
            )
        elif model_name in GPT5_MODEL_SERIES:
            # gpt5-low/medium/high → 実際の model_name は gpt5 に固定
            chat_response = await AzureOpenAI(
                chat_messages,
                model_name="gpt5",
                max_retries=2,
                timeout=230,
                raise_for_error=False,
                registry=LLM_REGISTRY,
                reasoning_effort=GPT5_MODEL_SERIES[model_name],
                verbosity="high",
            )
        elif model_name.startswith("gpt5"):
            # 他の gpt5 系モデルは low + high verbosity
            chat_response = await AzureOpenAI(
                chat_messages,
                model_name=model_name,
                max_retries=2,
                timeout=230,
                raise_for_error=False,
                registry=LLM_REGISTRY,
                reasoning_effort="low",
                verbosity="high",
            )
        else:
            chat_response = await AzureOpenAI(
                chat_messages,
                model_name=model_name,
                max_retries=2,
                timeout=230,
                raise_for_error=False,
                registry=LLM_REGISTRY,
            )
        # chat_response["choices"][0]["message"]["content"] += "<br><hr>" + "インターネット上の情報を検索して生成した回答が欲しい場合は「1」を入力してください。"
        return func.HttpResponse(json.dumps(chat_response))
    except Exception as e:
        logging.critical(f"LLM_CHAT: {e}")
        return func.HttpResponse(
            json.dumps(
                error_response(
                    "サーバーからの応答がありません。時間をおいてお試しください。"
                )
            )
        )


# LLM_GOOGLE
@app.route(route="llm/google", methods=("POST",))
async def llm_google(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("LLM_GOOGLE processed a request.")
    ########
    # init #
    ########

    async_http_client = AsyncHttpClient()
    # get json input
    req_json = req.get_json()
    sendFrom = req_json["from"]

    model_name = req_json.get("model", "gpt4.1")
    query_model_name = "gpt4.1-mini"

    max_query_num = 5  # クエリを生成する数は5
    max_search_results = 5  # 検索件数は5
    max_link_num = 3  # AOAIのリンク選択件数は3
    link_retry_limit = 1  # AOAIのリンク検索のリトライ回数は1

    messages = req_json["messages"]
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)

    # prompt
    cse_system_content = CSE_SYSTEM_CONTENT
    cse_system_content = cse_system_content.format(formatted_date=formatted_date())

    ui_system_content = ""
    if sendFrom == "web":
        ui_system_content = WEB_SYSTEM_CONTENT
    if sendFrom == "mail":
        ui_system_content = MAIL_SYSTEM_CONTENT
    if sendFrom == "teams":
        ui_system_content = TEAMS_SYSTEM_CONTENT
    cse_result_system_content = (
        BASE_SYSTEM_CONTENT + ui_system_content + CSE_RESULT_SYSTEM_CONTENT
    )
    cse_result_system_content = cse_result_system_content.format(
        formatted_date=formatted_date()
    )

    ########
    # main #
    ########

    # 検索queryの生成
    system_prompt = QUERY_GENERATION_SYSTEM_CONTENT.format(
        max_query_num=max_query_num, messages=messages
    )
    QueryList.model_fields["queries"].description = QueryList.model_fields[
        "queries"
    ].description.format(max_query_num=max_query_num)

    tools = [openai.pydantic_function_tool(QueryList)]

    _messages = deepcopy(messages)
    _messages.append({"role": "system", "content": system_prompt})

    google_input_response = await AzureOpenAI(
        messages=_messages,
        model_name=query_model_name,
        tools=tools,
        tool_choice="required",
        timeout=300,
        max_retries=2,
        raise_for_error=False,
        registry=LLM_REGISTRY,
    )

    args = google_input_response["choices"][0]["message"]["tool_calls"][0]["function"][
        "arguments"
    ]
    queries = json.loads(args).get("queries", [])
    logging.info(f"Generated queries: {queries}")

    if len(queries) == 0:
        logging.warning("Google: No query")
        google_error_response = error_response(
            "検索キーワードの作成に失敗しました。別の表現をお試しください。"
        )
        return func.HttpResponse(json.dumps(google_error_response))

    if len(queries) > max_query_num:
        logging.warning("queryを減らしました")
        queries = queries[:max_query_num]

    # 各クエリごとに5件ずつ検索して結果を集約
    cse_list = []
    seen_links = set()  # 重複チェック用のセット
    for query in queries:
        logging.info("query: " + query)
        try:
            search_result = await google_search(
                async_http_client, query, num=max_search_results
            )
        except Exception as e:
            logging.critical(f"Google API ERROR {e}, query: {query}")
            continue

        # エラーチェック（検索結果の先頭の要素を確認）
        if len(search_result) > 0 and search_result[0].get("title") == "ERROR":
            logging.critical(
                "Google API ERROR: " + search_result[0].get("snippet", "Unknown error")
            )
            continue

        # 重複するリンクを除外して追加
        for item in search_result:
            link = item.get("link")
            if link and link not in seen_links:
                cse_list.append(item)
                seen_links.add(link)
            else:
                logging.info(f"重複リンクをスキップ: {link}")

    logging.info(f"重複排除後の検索結果数: {len(cse_list)}")

    if len(cse_list) == 0:
        logging.critical(f"no cse_list")
        google_error_response = error_response("Googleでの検索に失敗しました。")
        return func.HttpResponse(json.dumps(google_error_response))

    # AOAIのリンク選択ループの前にデフォルト値を設定
    selected_cse_list = cse_list[:max_link_num]

    # リンク数が十分にある場合のみAOAIのリンク選択を実行
    if len(cse_list) > max_link_num:
        # ツール・プロンプトの設定
        link_tools = [openai.pydantic_function_tool(LinkList)]
        LinkList.model_fields["selected_links"].description = LinkList.model_fields[
            "selected_links"
        ].description.format(max_link_num=max_link_num)
        prompt_for_link_selection = LINK_SELECTION_SYSTEM_CONTENT.format(
            max_link_num=max_link_num,
            messages=messages,
            candidates=json.dumps(cse_list, ensure_ascii=False, indent=2),
        )
        message_for_link_selection = [
            {"role": "system", "content": prompt_for_link_selection}
        ]

        for attempt in range(link_retry_limit + 1):
            try:
                link_selection_response = await AzureOpenAI(
                    messages=message_for_link_selection,
                    model_name=query_model_name,
                    tools=link_tools,
                    tool_choice="required",
                    timeout=300,
                    max_retries=1,
                    raise_for_error=False,
                    registry=LLM_REGISTRY,
                )
                logging.info(
                    f"[Attempt {attempt+1}] link_selection_response: {link_selection_response}"
                )

                # AOAIの応答から選ばれたリンク(selected_links)を抽出
                calls = link_selection_response["choices"][0]["message"].get(
                    "tool_calls", []
                )
                if not calls:
                    logging.warning("tool_callsが空です。リトライします。")
                    continue
                call = calls[0]
                if call["function"]["name"] == "LinkList":
                    try:
                        args = json.loads(call["function"]["arguments"])
                        selected_links = args.get("selected_links", [])
                    except Exception as e:
                        logging.warning(
                            f"JSON パース中にエラーが発生しました: {e}。リトライします。"
                        )
                        continue

                    # リスト化処理の前に選択されたリンクの数が正しいかを確認
                    if len(selected_links) != max_link_num:
                        logging.warning(
                            f"選択されたリンクは{len(selected_links)}件でした。リトライします。"
                        )
                        continue
                    # 選択したリンクとcse_listのリンクが一致し、リンク数も一致すれば終了
                    selected_cse_list = [
                        item for item in cse_list if item.get("link") in selected_links
                    ]
                    if len(selected_cse_list) == max_link_num:
                        break

                    # 検索件数が一致しない場合はリトライ
                    logging.warning(
                        f"選択リンク数 {len(selected_cse_list)} 件 リンク選択をリトライします。"
                    )
                else:
                    logging.warning(
                        f"予期しないfunction名: {call['function']['name']}. リトライします。"
                    )
                    continue

            except Exception as e:
                logging.warning(
                    f"リンク選択レスポンスの構造解析中に予期せぬエラーが発生しました: {e}。リトライします。"
                )
                continue

    if model_name == "o4-mini" or model_name.startswith("claude-"):
        max_length = 30_000  # 選択したリンク３件分の本文の総文字数が90,000文字を超えないように設定

    elif model_name == "gpt5":
        max_length = 40_000  # 選択したリンク３件分の本文の総文字数が120,000文字を超えないように設定

    else:  # gpt-4.1の場合
        max_length = 700_000

    # 選択したリンクにアクセスして本文を取得
    updated_results = await updateSearchResults(
        selected_cse_list, max_length=max_length
    )

    # 取得した本文をcse_listのsnippetに上書き更新
    link_to_snippet_map = {}
    for item in updated_results:
        key = item["link"]
        value = item["snippet"]
        link_to_snippet_map[key] = value
    for entry in cse_list:
        link = entry["link"]
        if link in link_to_snippet_map:
            entry["snippet"] = link_to_snippet_map[link]

    # スニペット、タイトル、リンクの組み立て
    cse_snippet = "<br>".join(snippet["snippet"] for snippet in cse_list)
    logging.info("snippet:<br>" + cse_snippet)

    cse_title = "<br>".join(snippet["title"] for snippet in cse_list)
    logging.info("title:<br>" + cse_title)

    cse_title_link = "【参照元URL】"
    for snippet in cse_list[0:3]:
        cse_title_link += "<br><a href='{url}' target='_blank'>{title}</a>".format(
            url=snippet["link"], title=snippet["title"]
        )

    cse_pre = f"Googleで{queries}を検索した結果、以下の回答が得られました。<br>"

    # 最終回答の生成
    cse_result_messages = change_system_content(
        messages, system_content=cse_result_system_content
    )
    cse_assistant_message = {"role": "user", "content": cse_pre + cse_snippet}
    if sendFrom in [
        "teams",
    ]:
        cse_assistant_message["content"] += "\nHTML形式で回答してください。"
    cse_result_messages.append(cse_assistant_message)

    try:
        if model_name.startswith("claude-"):
            # Claude モデルの場合
            google_response = await ClaudeGenerate(
                cse_result_messages,
                temperature=0,
                model_name=model_name,
                max_retries=2,
                raise_for_error=False,
                registry=LLM_REGISTRY,
            )
        elif model_name in GPT5_MODEL_SERIES:
            # gpt5-low/medium/high → 実際の model_name は gpt5 に固定
            google_response = await AzureOpenAI(
                cse_result_messages,
                temperature=0,
                model_name="gpt5",
                max_retries=2,
                raise_for_error=False,
                registry=LLM_REGISTRY,
                reasoning_effort=GPT5_MODEL_SERIES[model_name],
                verbosity="high",
            )
        elif model_name.startswith("gpt5"):
            # 他の gpt5 系モデルは low + high verbosity
            google_response = await AzureOpenAI(
                cse_result_messages,
                temperature=0,
                model_name=model_name,
                max_retries=2,
                raise_for_error=False,
                registry=LLM_REGISTRY,
                reasoning_effort="low",
                verbosity="high",
            )
        else:
            google_response = await AzureOpenAI(
                cse_result_messages,
                temperature=0,
                model_name=model_name,
                max_retries=2,
                raise_for_error=False,
                registry=LLM_REGISTRY,
            )
        google_response_text = google_response["choices"][0]["message"]["content"]
        logging.info("Google assistant: " + google_response_text)
        google_response["choices"][0]["message"]["content"] += (
            "<br><hr>" + cse_title_link
        )

        # suffix
        return func.HttpResponse(json.dumps(google_response))
    except:
        logging.critical("Google AzureOpenAI: no response")
        google_error_response = error_response(
            "サーバーからの応答がありません。時間をおいてお試しください。"
        )
        return func.HttpResponse(json.dumps(google_error_response))


# LLM_GEMINI
@app.route(route="llm/gemini", methods=("POST",))
async def llm_gemini(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("LLM_GEMINI processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    sendFrom = req_json["from"]
    model_name = req_json["model"]
    logging.info("Gemini model: " + model_name)
    messages = req_json["messages"]
    messages = messages[-15:]  # 直近のやり取り15件に絞ります。
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)

    ## ステップ2. ユーザーとのやり取りとシステムメッセージをGeminiに渡します。
    ui_system_content = ""
    if sendFrom == "web":
        ui_system_content = WEB_SYSTEM_CONTENT
    if sendFrom == "mail":
        ui_system_content = MAIL_SYSTEM_CONTENT
    if sendFrom == "teams":
        ui_system_content = TEAMS_SYSTEM_CONTENT
    gemini_prompt = (
        BASE_SYSTEM_CONTENT + ui_system_content + SEARCH_GROUNDING_PROMPT
    )  # Geminiのシステムプロンプトは元々のI-Colleagueのプロンプトを使用
    gemini_prompt = gemini_prompt.format(formatted_date=formatted_date())
    gemini_messages = [
        {"role": "system", "content": gemini_prompt}
    ] + messages  # システムプロンプトとユーザーとのやり取りをGeminiに渡します。

    ## ステップ3. Geminiから回答結果を取得し、参照元などを必要に応じて整形して返します。
    try:
        gemini_response = await GeminiGenerate(
            gemini_messages,
            temperature=0,
            model_name=model_name,
            raise_for_error=False,
            grounding=True,  # グラウンディングを有効化
            auth_type="json_key",  # JSONキー認証を仕様
            registry=LLM_REGISTRY,
            labels=GEMINI_DEFAULT_LABELS,
        )
        logging.info("Gemini processing completed successfully")
        gemini_answer = gemini_response["choices"][0]["message"][
            "content"
        ]  ## Geminiの回答
        logging.info("Gemini answer: " + gemini_answer)
        google_search_info = gemini_response.get(
            "google_search_info", {}
        )  ## GeminiがGoogle検索を行った内容
        if (
            google_search_info
        ):  # もしGeminiがGoogle検索を行った場合、検索結果を生成して、回答に追加します。
            # grounding_chunks が None の場合に備えてフォールバック
            chunks = google_search_info.get("grounding_chunks") or []
            if not isinstance(chunks, list):
                chunks = []
            reference_text = "".join(
                f"<br><a href='{chunk.get('web', {}).get('uri', '#')}' target='_blank'>"
                f"{chunk.get('web', {}).get('title', 'No Title')}</a>"
                for chunk in chunks
                if isinstance(chunk, dict)
                and chunk.get("web")
                and chunk["web"].get("uri")
            )
            if reference_text:  # 参照したurlがある場合、整形して回答に追加します
                gemini_answer += (
                    GEMINI_IGNORE_SIGN + f"<br>【参照元URL】{reference_text}"
                )
            # <a>タグに_blankを追加します。
            search_entry_point = google_search_info.get("search_entry_point") or {}
            rendered_content = search_entry_point.get("rendered_content", "")
            if rendered_content:
                # <a>タグに target="_blank" を追加
                search_response_result = re.sub(
                    r"(<a\b(?![^>]*\btarget=))", r'\1 target="_blank"', rendered_content
                )
                # .container の style に width: 100% を追加
                search_response_result = re.sub(
                    r"(\.container\s*\{)", r"\1 width: 100%;", search_response_result
                )
                gemini_response["choices"][0]["message"]["content"] = (
                    gemini_answer + search_response_result
                )
            else:
                gemini_response["choices"][0]["message"]["content"] = gemini_answer
        logging.info("Gemini successfully processed the response")
        return func.HttpResponse(json.dumps(gemini_response))
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"GeminiGenerate ERROR: {str(e)}, {str(tb)}")
        # logging.critical("GeminiGenerate ERROR")
        return func.HttpResponse(
            json.dumps(
                error_response(
                    "サーバーからの応答がありません。時間をおいてお試しください。"
                )
            )
        )


# azure-openai-documents-search wrapper
@app.route(route="llm/docs", methods=("POST",))
async def llm_docs(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("docs processed a request.")
    ########
    # init #
    ########

    # get json input
    req_json = req.get_json()
    messages = req_json["messages"]
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)

    upn = req_json["upn"]

    DOCUMENTS_SEARCH_API_KEY = os.environ.get("DOCUMENTS_SEARCH_API_KEY")
    DOCUMENTS_SEARCH_URL = os.environ.get("DOCUMENTS_SEARCH_URL")

    # 辞書を参照してpromptの置換
    prompt = user_input

    for key, word_list in replace_dict.items():
        for word in word_list:
            prompt = prompt.replace(word, key)
    logging.info("prompt: " + prompt)

    history_list = []
    # messages の末尾は最新入力なので、それを含めずそれ以前を履歴にする
    for i in range(len(messages) - 2):
        if messages[i]["role"] == "user" and messages[i + 1]["role"] == "assistant":
            user_text = next(
                (c["text"] for c in messages[i]["content"] if c["type"] == "text"), ""
            )
            assistant_text = next(
                (c["text"] for c in messages[i + 1]["content"] if c["type"] == "text"),
                "",
            )
            history_list.append({"user": user_text, "assistant": assistant_text})

    # リクエストボディのデータを準備
    json_data = {
        "upn": upn,
        "prompt": prompt,
        "history": history_list,
        "overrides": VARIABLE_LIST[ENVIRONMENT_SELECTED]["llm_docs_overrides"],
    }

    # お知らせ機能
    info_suffix = "<br><hr>" + generate_info_suffix(
        user_input=user_input,
        file_path="./data/検索連動型お知らせ機能_設定内容_20231219.xlsm",
        sheet_name="「検索連動型お知らせ機能」設定内容（本番）",
    )

    ########
    # main #
    ########
    # Documents Search API呼び出し
    api_name = "Documents search"
    try:
        cs_response = await http_post(
            json_data=json_data,
            url=DOCUMENTS_SEARCH_URL,
            api_key=DOCUMENTS_SEARCH_API_KEY,
            process_name=api_name,
        )
        # responseの整形
        ans = cs_response["answer"]
        logging.info("cs answer: " + ans)
        ans += info_suffix
    except Exception as e:
        error = api_name + ": " + str(e)
        logging.critical(error)

        return func.HttpResponse(
            json.dumps(
                error_response(
                    "社内情報との接続が出来ませんでした。時間をおいてお試しください。"
                    + info_suffix
                )
            )
        )

    try:  # 資料が存在する場合
        res_suffix = "<br><hr>【参照元URL】"
        # 綴り間違えて実装されてます
        if "datasuorce" in cs_response.keys():
            data_source = cs_response["datasuorce"]
        else:
            data_source = cs_response["datasource"]
        # 各種パラメータの取得
        counter = 0
        for num in range(len(data_source)):
            doc_file_path = data_source[num]["filepath"]  # "OO編_主催者用.pdf"
            doc_url = data_source[num]["url"]

            if doc_url in res_suffix:  # 重複する場合は追加しない
                continue
            counter += 1
            res_suffix += "<br><a href='{url}' class='custom-link' target='_blank'>{title}</a>".format(
                url=doc_url, title=doc_file_path
            )
            if counter == 3:  # 3つの資料のみ表示
                break
        # 締めの文（ループの外に配置）
        res_suffix += "<br><hr>" + "正確な情報については、上記URLをご参照ください。"
    except Exception as e:  # 資料が存在しない場合
        logging.warning(f"datasource: {e}")
        res_suffix = "<br><hr>" + "資料を見つけられませんでした。"

    ans += res_suffix

    res = {
        "id": "***********",
        "object": "LLM_DOCS",
        "created": 1234567890,
        "model": "gpt-4-o",
        "choices": [
            {
                "index": 0,
                "finish_reason": "stop",
                "message": {"role": "assistant", "content": ans},
            }
        ],
        "usage": {"completion_tokens": 0, "prompt_tokens": 0, "total_tokens": 0},
        "ans": cs_response["answer"],
        "datasource": data_source,
    }

    return func.HttpResponse(json.dumps(res))


# azure-openai-documents-search wrapper
@app.route(route="llm/enq", methods=("POST",))
async def llm_enq(req: func.HttpRequest) -> func.HttpResponse:
    """
    エネ化問い合わせ集約検証用
    3つのAPIを並列で呼び出す。
    - 全社イントラ検索
    - エネ化イントラ検索
    - エネ化問い合わせDB検索
    """
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    # get json input
    req_json = req.get_json()
    upn = req_json["upn"]
    formatted_messages = req_json["messages"]
    # ファイルアップロード機能の処理の追加

    user_input = next(
        (
            content["text"]
            for content in formatted_messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info(f"User input: {user_input}")

    ########
    # main #
    ########
    # 非同期で並列処理
    enq_api_call = EnqAPICall(
        req_json=req_json, authority_verification=AuthorityVerification
    )
    responses = await enq_api_call.call_apis()

    unanswerable_phrases = [
        "回答できません",
        "わかりません",
        "申し訳ありませんが",
        "情報が不足しています",
        "お答えできません",
        "与えられた情報からは",
    ]
    combined_answers = {}
    enq_departments = set()
    for response in responses:
        logging.info(f"mode: {response['mode']}, response: {response}")

        mode = response["mode"]
        ans = response["answer"]
        reference_info = f"【参照元】<br>"
        urls = []
        seen_documents = set()

        if response.get("datasource"):
            for source in response["datasource"]:
                if mode == "enquiry" and source.get("type") == "msg":
                    enq_departments.add(source["department"])
                if source["url"]:
                    document_name = source["title"]
                    if document_name not in seen_documents:
                        url = source["url"]
                        urls.append(
                            f"<a href='{url}' target='_blank'>{document_name}</a>"
                            if url
                            else f"{document_name}"
                        )
                        seen_documents.add(document_name)

            if urls:
                reference_info += "<br>".join(urls)
            else:
                reference_info = ""
        else:
            reference_info = ""
        logging.info(f"Enquiry ans: {ans}")

        role_info = f"※ より詳細な情報が必要な場合は以下の役割表をご確認の上、担当者にご連絡ください。<br><a href='https://itochucorp.sharepoint.com/sites/company-eneka/DocLib003/Forms/View03.aspx?id=%2Fsites%2Fcompany%2Deneka%2FDocLib003' target='_blank'>エネルギー・化学品カンパニーイントラネット - 資料 - 40.役割表</a>"
        if response["mode"] == "enquiry":
            if any(phrase in ans for phrase in unanswerable_phrases):
                reference_info = role_info
            elif enq_departments:
                departments_str = "、".join(list(enq_departments))
                ans += f"<br><br>※ 過去に{departments_str}が本件に関連する問合せに回答しています。"
                reference_info = f"{role_info}<br><br>{reference_info}"
            else:
                reference_info = f"{role_info}<br><br>{reference_info}"
            combined_answers[response["name"]] = {
                "answer": ans,
                "reference": reference_info,
            }
        else:
            combined_answers[response["name"]] = {
                "answer": ans,
                "reference": reference_info,
            }

    # full_textの作成
    full_text = []
    for source, data in combined_answers.items():
        ans = data["answer"]
        full_text.append(f'<p style="font-weight: bold;">■ {source}</p><md>{ans}</md>')
        if data["reference"]:
            full_text.append(data["reference"])

    text_response = "<br><br>".join(full_text)

    # お知らせ機能
    info_suffix = generate_info_suffix(
        user_input=user_input,
        file_path="./data/検索連動型お知らせ機能_設定内容_20231219.xlsm",
        sheet_name="「検索連動型お知らせ機能」設定内容（本番）",
    )

    if info_suffix:
        full_text_response = (
            "<br><hr>" + info_suffix + "<br><hr><br><br>" + text_response
        )
    else:
        full_text_response = text_response
    res = {
        "id": "***********",
        "object": "LLM_ENQUIRY",
        "created": 1234567890,
        "model": "gpt-4-o",
        "blobs": [],
        "choices": [
            {
                "index": 0,
                "finish_reason": "stop",
                "message": {
                    "role": "assistant",
                    "content": full_text_response,
                },
            }
        ],
        "usage": {"completion_tokens": 0, "prompt_tokens": 0, "total_tokens": 0},
    }

    return func.HttpResponse(json.dumps(res))


# LLM_MINUTES
@app.route(route="llm/minutes", methods=("POST",))
async def llm_minutes(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("LLM_MINUTES processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    messages = req_json["messages"]
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)

    mail = req_json.get("mail")

    if len(user_input) < 50:
        logging.warning(f"too short input: {len(user_input)}")
        return func.HttpResponse(json.dumps(error_response("文字数が足りません。")))

    # 仮置き
    # gpt4.1
    model_name = "o4-mini"
    max_tokens = NON_CHAT_REGISTRY.models[model_name].max_tokens
    model_max_token = max_tokens["input"]

    # ループ制御
    loop_count = 0
    start_num = 0
    len_buffer = 500
    len_text = max_len_text = model_max_token - max_tokens["output"] - len_buffer * 2

    # prompt
    minutes_system_content = MINUTES_SYSTEM_CONTENT
    minutes_system_content = minutes_system_content.format(
        formatted_date=formatted_date()
    )

    ########
    # main #
    ########
    minutes_messages = change_system_content(messages, minutes_system_content)
    user_prefix = "残りの会話ログです。先程の議事録を踏まえてあらためて完成版の議事録を作成してください。\n"
    try:
        while True:
            # loop init
            end_num = start_num + len_text + len_buffer
            end_num = min(end_num, len(user_input[start_num:]))

            minutes_messages = []
            minutes_messages.append(
                {"role": "system", "content": minutes_system_content}
            )
            if loop_count >= 1:
                minutes_messages.append(
                    {"role": "assistant", "content": response_content}
                )
                minutes_messages.append(
                    {
                        "role": "user",
                        "content": user_prefix + user_input[start_num:end_num],
                    }
                )
            else:
                minutes_messages.append(
                    {"role": "user", "content": user_input[start_num:end_num]}
                )

            # token数を超過していないかのチェック
            tokens = 0
            for message in minutes_messages:
                tokens += check_token(message["content"])
            logging.warning(f"tokens: {tokens}")
            if model_max_token < tokens + max_tokens["output"] + 1:
                logging.warning(f"minutes: too many token {tokens}")
                len_text -= 1000
                continue

            minutes_response = await AzureOpenAI(
                minutes_messages,
                temperature=0,
                model_name=model_name,
                max_retries=2,
                timeout=230,
                raise_for_error=False,
                registry=NON_CHAT_REGISTRY,
            )
            response_content = minutes_response["choices"][0]["message"]["content"]
            logging.info(f"loop count: {loop_count}, minutes: {response_content}")

            if end_num == len(user_input[start_num:]):
                break
            else:
                loop_count += 1
                start_num += len_text
                len_text = max_len_text

    except Exception as e:
        # エラーメッセージをログファイルに記録
        logging.error("LLM_MINUTES: " + f"Error occurred: {e}")
        return func.HttpResponse(
            json.dumps(
                error_response(
                    "サーバーとの接続が出来ませんでした。時間をおいてお試しください。"
                )
            )
        )

    response_content = minutes_response["choices"][0]["message"]["content"]
    # 英語版議事録の作成 # リファクタリング予定
    json_data = {
        "language": "en",
        "messages": [
            {"role": "user", "content": [{"type": "text", "text": response_content}]}
        ],
    }
    try:
        translation_response = await http_post(
            json_data=json_data,
            url=os.environ.get("LLM_TRANSLATION_URL"),
            api_key=os.environ.get("LLM_TRANSLATION_API_KEY"),
            process_name="LLM_TRANSLATION",
        )
        minutes_response["choices"][0]["message"]["content"] += (
            "\n<hr>\n" + translation_response["choices"][0]["message"]["content"]
        )
    except Exception as e:
        error = "LLM_TRANSLATION" + ": " + str(e)
        logging.critical(error)

    # send_mail
    if mail != None:
        try:
            attachment_list = [
                {
                    "file_name": "minutes.txt",
                    "file_content": minutes_response["choices"][0]["message"][
                        "content"
                    ].encode("utf-8"),
                },
                {
                    "file_name": "transcribe.txt",
                    "file_content": user_input.encode("utf-8"),
                },
            ]
            send_email(
                subject="【I-Colleague】文字起こし結果を送付いたします。",
                content="会議の文字起こし、議事録データを送信いたします。\nこのメールに添付ファイル付きで返信していただくと、内容についてメールにて回答できます。（例：「添付ファイルの内容を要点にまとめて」）",
                send_to=[
                    mail,
                ],
                attachment_list=attachment_list,
            )

        except Exception as e:
            logging.warning(f"cannot send minutes: {e}")
            pass
    return func.HttpResponse(json.dumps(minutes_response))


# LLM_TRANSLATION
@app.route(route="llm/translation", methods=("POST",))
async def llm_translation(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("LLM_TRANSLATION processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    messages = req_json["messages"]
    user_input = next(
        (
            content["text"]
            for content in messages[-1]["content"]
            if content["type"] == "text"
        ),
        "",
    )
    logging.info("User input:" + user_input)
    if "language" in req_json.keys():
        language = req_json["language"]
    else:
        language = "ja"

    for key, item in enumerate(language_dict):
        if key == language:
            language = item
    logging.info(f"language: {language}")

    model_name = "gpt4.1-mini"

    ########
    # main #
    ########
    translation_system_content = TRANSLATION_SYSTEM_CONTENT
    translation_system_content = translation_system_content.format(
        language=language, user_input=user_input
    )
    translation_messages = [{"role": "system", "content": translation_system_content}]
    try:
        translation_response = await AzureOpenAI(
            translation_messages,
            temperature=0,
            model_name=model_name,
            max_retries=2,
            timeout=230,
            raise_for_error=False,
            registry=LLM_REGISTRY,
        )
        logging.info(
            f'translation: {translation_response["choices"][0]["message"]["content"]}'
        )
        return func.HttpResponse(json.dumps(translation_response))
    except Exception as e:
        logging.critical(f"LLM_TRANSLATION: {e}")
        return func.HttpResponse(
            json.dumps(
                error_response(
                    "サーバーからの応答がありません。時間をおいてお試しください。"
                )
            )
        )


# #########
# # genie #
# #########
# # GENIE_file???
# @app.route(route="genie/vision", methods=("POST",))
# def genie_vision(req: func.HttpRequest) -> func.HttpResponse:
# # data = {
# #     "upn": "<upn>", # オプション
# #     "messsages" :[ # 必須
# #         { "role": "system", "content": "You are a helpful assistant." },
# #         { "role": "user", "content": [
# #             { # 必須
# #                 "type": "file",
# #                 "name": "<ファイルの名前>", # sample.pdf
# #                 "data": "<base64エンコーディングされたデータ>"
# #             },
# #             { # オプション
# #                 "type": "text", # "text" or "file"
# #                 "text": "このファイルの背景として以下のような情報があります、、、、、"
# #             }
# #         ] }
# #     ]
# # }
# ########
# # init #
# ########
#     req_json = req.get_json()
#     if "upn" in req_json.keys():
#         upn = req_json["upn"]
#         logging.info("upn: " + upn)
#     else:
#         logging.critical("no upn")
#     try:
#         vision_messages = format_messages(messages=req_json["messages"], max_tokens=127_000)
#     except Exception as e:
#         logging.critical(f"error in format_messages: {e}")
#         return func.HttpResponse(json.dumps(error_response(f"error in format_messages: {e}")))
#         return func.HttpResponse(json.dumps(error_response(f"ファイルを読み込めませんでした。")))
# ########
# # main #
# ########
#     # vision_messages = [] + vision_messages
#     try:
#         vision_response = await AzureOpenAI(vision_messages, temperature=0, model_name=4)
#         return func.HttpResponse(json.dumps(vision_response))
#     except Exception as e:
#         logging.critical(f"GENIE_VISION: {e}")
#         return func.HttpResponse(json.dumps(error_response("サーバーからの応答がありません。時間をおいてお試しください。")))


# PA
@app.route(route="genie/powerapps", methods=("POST",))
async def powerapps_dev(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("PA processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    messages = req_json["messages"]
    user_input = req_json["messages"][-1]["content"]
    # if "logit_bias" in req_json.keys():
    #     logit_bias = req_json["logit_bias"]
    logging.info(f"User input: {user_input}")

    # モデルの固定
    req_json["model_name"] = "gpt4.1"

    ########
    # main #
    ########
    try:
        pa_response = await AzureOpenAI(**req_json, registry=LLM_REGISTRY)
        return func.HttpResponse(json.dumps(pa_response))
    except Exception as e:
        logging.critical(f"GENIE_PA: {e}")
        return func.HttpResponse(
            json.dumps(
                error_response(
                    "サーバーからの応答がありません。時間をおいてお試しください。"
                )
            )
        )


# crm_response
@app.route(route="genie/crm/response", methods=("POST",))
async def crm_create_response(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("CRM_CREATE_RESPONSE processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    mode = req_json["mode"]  # aoai, crm
    try:
        user_input = req_json["messages"][-1]["content"]
    except Exception as e:
        logging.debug(f"error in loading user_input: {e}")
        user_input = ""
    logging.info("User input:" + user_input)

    req_json["temperature"] = 0
    req_json["model_name"] = "gpt4.1"

    if mode == "crm":
        logging.debug("crm")
        # debug
        options = {}
        if "options" in req_json.keys():
            options = req_json["options"]
        logging.debug(f"{options}")

    else:
        if mode == "aoai":
            try:
                res_json = await AzureOpenAI(**req_json, registry=LLM_REGISTRY)
            except Exception as e:
                logging.critical(f"GENIE_CRM_RESPONSE: {e}")
                res_json = json.dumps(
                    error_response(
                        "サーバーからの応答がありません。時間をおいてお試しください。"
                    )
                )
        else:
            logging.critical(f"無効なモード: {mode}")
            res_json = error_response(f"無効なモード: {mode}")
        return func.HttpResponse(json.dumps(res_json))

    # print(crm_df)

    ########
    # main #
    ########
    # インプットから検索のスケジューリングの作成
    # query = conv2query(req_json)
    # logging.critical(str(query))
    # 仮置き
    # query = user_input
    # crm_data = await vector_search(query)
    # logging.critical(str(crm_data))

    # 検索の実行
    crm_df = csv_search(**options)
    if len(crm_df) == 0:
        res_json = error_response(
            f"該当データがありません。条件を変えて検索してください。"
        )
        return func.HttpResponse(json.dumps(res_json))
    # 検索結果の整形
    crm_data = crm_df.to_markdown()
    if len(crm_data) > 15_000:
        res_json = error_response(
            f"該当データが多すぎます。条件を絞って検索してください。"
        )
        logging.warning(f"too much results: {len(crm_data)}")
        return func.HttpResponse(json.dumps(res_json))

    # messagesの作成
    crm_system_content = CRM_SYSTEM_CONTENT
    crm_system_content = crm_system_content.format(crm_data=crm_data)

    crm_user_content = CRM_USER_CONTENT
    crm_user_content = crm_user_content.format(
        担当者=options["担当者"],
        start=options["対象期間"]["start"],
        end=options["対象期間"]["end"],
        対象=options["対象"],
    )
    req_json["messages"] = [
        {"role": "system", "content": crm_system_content},
        {"role": "user", "content": crm_user_content},
    ]

    # 最終回答の生成
    try:
        crm_response = await AzureOpenAI(**req_json, registry=LLM_REGISTRY)
    except Exception as e:
        logging.critical(f"GENIE_CRM_RESPONSE: {e}")
        crm_response = error_response(
            "サーバーからの応答がありません。時間をおいてお試しください。"
        )
    return func.HttpResponse(json.dumps(crm_response))


# crm_create_db
@app.route(route="genie/crm/create_db", methods=("POST",))
async def crm_create_db(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("CRM_CREATE_DB processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    res_json = {"status": "failure"}

    ########
    # main #
    ########
    # データの取り込み
    try:
        encoded_csv = req_json["csv"]
        csv_data = base64.b64decode(encoded_csv)
        logging.debug(f"csv_data[:100]: {csv_data[:100]}")
    except Exception as e:
        logging.critical(f"CANNOT decode csv: {e}")
        res_json["error"] = f"Error in decoding data: {e}"
        func.HttpResponse(json.dumps(res_json))

    # データのblobへのアップロード
    status_flag = upload_blob(
        file_name="food_co/crm.csv",
        file_content=csv_data,
        container_name="crm",
        overwrite=True,
    )
    if status_flag == False:
        res_json["error"] = f"CANNOT access blob"
        func.HttpResponse(json.dumps(res_json))

    res_json["status"] = "success"
    return func.HttpResponse(json.dumps(res_json))


# ML_WHISPER


### input
# data = {
#     "upn": <upn>,
#     "mail": <mail>,
#     "inputs": {
#         "url": <url>,
#         "container_name": <container>,
#         "content": ["<Base64エンコードされたデータ>"],
#         "file_name": ["<ファイル名(sample.pdf)>"],
#     }
# }
@app.route(route="genie/upload", methods=("POST",))
async def upload_file(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("UPLOAD_FILE processed a request.")
    ########
    # init #
    ########
    # get json input
    req_json = req.get_json()
    json_response = {"status": 500}
    upn = req_json["upn"]
    logging.info("upn: " + upn)

    try:
        mail = req_json["mail"]
        logging.info("mail: " + mail)
    except Exception as e:
        logging.warning(f"no mail: {e}")

    # inputs
    inputs = req_json["inputs"]
    url = inputs["url"]
    container_name = inputs["container_name"]

    # Base64デコードしてバイナリデータに変換
    encoded_data = inputs["content"][0]
    binary_data = base64.b64decode(encoded_data)

    # file_name
    file_name = inputs["file_name"][0]
    logging.debug(f"uploaded: {file_name}")

    ########
    # main #
    ########
    ## upload blob
    try:
        upload_blob(
            file_name=file_name, file_content=binary_data, container_name=container_name
        )
    except Exception as e:
        logging.critical(f"CANNOT UPLOAD FILE: {e}")
        json_response["error"] = f"CANNOT UPLOAD FILE: {e}"
        return func.HttpResponse(json.dumps(json_response))

    # blobのurlを渡す
    ## durableのraise event
    headers = {
        "Content-Type": "application/json",
        # "x-functions-key":api_key
    }
    data = {"blob_name": file_name}

    # raise event
    data_encode = json.dumps(data)
    response = requests.post(url, headers=headers, data=data_encode)

    logging.info(f"code: {response.status_code}")
    json_response["status"] = 202
    json_response["message"] = "accept"
    return func.HttpResponse(json.dumps(json_response))


### input
# data = {
#     "inputs": {
#         "audio": ["<Base64エンコードされた音声データ>"],
#         "language": [], # optional 入力される言語の指定 通常は空で 固定する場合は指定 必要なし
#         "suffix":[".mp3"] # optional 音声ファイルの形式
#     "outputs": {
#         "language": ["ja"], # optional 翻訳を行う場合、出力する言語の設定
#     }
# }


@app.route(route="genie/whisper", methods=("POST",))
async def whisper(req: func.HttpRequest) -> func.HttpResponse:
    # logging.info('GENIE_WHISPER processed a request.')
    ########
    # init #
    ########
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    title = "No title"

    # apiの呼び出し
    history_base_url = os.environ.get("HISTORY_API_URL")
    history_api_key = os.environ.get("HISTORY_API_KEY")

    # get json input
    req_json = req.get_json()
    # body = str.encode(json.dumps(req_json))
    session_id = req_json.get("session_id")
    logging.info(f"session_id: {session_id}")
    try:
        upn = req_json["upn"]
        logging.info("upn: " + upn)
    except Exception as e:
        logging.warning(f"no upn: {e}")

    try:
        sendFrom = req_json["from"]
        assert sendFrom in ["web", "teams", "mail", "agent"]
    except Exception as e:
        req_json["from"] = sendFrom = "web"

    # request bodyの初期化
    whisper_body = {}

    # デフォルトは指定なし（自動検出）
    language_code = ""

    # 入力言語のフィルター処理
    inputs = req_json["inputs"]
    if "language" in inputs.keys():
        in_language = inputs["language"]
        in_language = [
            language for language in in_language if language in language_dict.keys()
        ]
        inputs["language"] = in_language
        # フィルター後に言語があれば設定
        if in_language:
            language_code = in_language[0]

    # プロンプトを取得
    selected_transcribe_prompt = transcribe_prompt_dict.get(language_code, "")

    # mlでは未使用
    if "outputs" in req_json.keys():
        outputs = req_json["outputs"]
        if "language" in outputs.keys():
            out_language = outputs["language"][0]
            # logging.info(f"translate to {out_language}")
        # whisper_body["outputs"] = outputs

    ########
    # main #
    ########

    client = AsyncAzureOpenAI(
        api_version=GPT_API_VERSION,
        azure_endpoint=GPT4O_TRANSCRIBE_API_ENDPOINT,
        api_key=GPT4O_TRANSCRIBE_API_KEY,
    )

    # content
    encoded_audio = inputs["audio"][0]
    binary_audio = base64.b64decode(encoded_audio)

    try:
        # BytesIOを使用して音声データをメモリ内ファイルとして扱う
        with io.BytesIO(binary_audio) as audio_buffer:
            audio_buffer.name = "audio.mp3"

            # AOAI（transcribe）での非同期文字起こし実行
            transcription = await client.audio.transcriptions.create(
                file=audio_buffer,
                model=GPT4O_TRANSCRIBE_DEPLOYMENT_NAME,
                prompt=selected_transcribe_prompt,
                temperature=whisper_options["values"].get("temperature", 0.0),
                language=language_code,
                response_format="json",
            )

        decoded_response = transcription

    except Exception as e:
        tb = traceback.format_exc()
        logging.critical(f"AOAI Error: {str(e)}, {str(tb)}")
        error_json = {"error": "AOAI ERROR"}
        return func.HttpResponse(json.dumps(error_json))

    finally:
        # 現在の日本時間を取得
        japan_timezone = ZoneInfo("Asia/Tokyo")
        current_time_japan = datetime.now(japan_timezone)
        formatted_time = current_time_japan.strftime("%Y-%m-%d/%H:%M:%S")

        file_name = f"{upn}/{formatted_time}.mp3"

        # upload blob
        upload_blob(
            file_name=file_name, file_content=binary_audio, container_name="audio-data"
        )

    # 文字起こしの結果の後処理
    text = decoded_response.text

    logging.info(f"transcribe text: {text}")

    # 重複の除去
    text = remove_repeated_phrases(text)

    # initial_promptの悪影響の除去
    for i in range(len(replace_list)):
        replace_text = replace_list[i]
        text = text.replace(replace_text, "")

    if text == "- " or text == "。" or len(text) <= 2:
        text = ""

    # responseの用意
    json_response = {"text": text}

    # 翻訳の必要性の確認
    translation_response_text = None

    if "out_language" in locals() and text != "":
        translation_message = {
            "role": "user",
            "content": [{"type": "text", "text": text}],
        }
        json_data = {
            "language": out_language,
            "messages": [translation_message],
        }

        try:
            translation_response = await http_post(
                json_data=json_data,
                url=os.environ.get("LLM_TRANSLATION_URL"),
                api_key=os.environ.get("LLM_TRANSLATION_API_KEY"),
                process_name="LLM_TRANSLATION",
            )
            translation_response_text = translation_response["choices"][0]["message"][
                "content"
            ]
            translation_response_text = translation_response_text.strip("'\"")

            # logging.info("translation response: " + translation_response_text)
            json_response["translation"] = (
                '<div style="color: blue; ">' + translation_response_text + "</div>"
            )
        except Exception as e:
            error = "LLM_TRANSLATION" + ": " + str(e)
            logging.critical(error)

    # 汎用履歴への登録
    if session_id != None:
        bilingual_text = text + (
            "\n" + "*" * 20 + "\n" + translation_response_text
            if translation_response_text
            else ""
        )

        request_data = {
            "items": [
                {
                    "upn": upn,
                    "content": [{"type": "blob", "name": file_name}],
                    "role": "user",
                    "submode": "web",
                    "model": "gpt-4o-transcribe",
                    "from": sendFrom,
                    "sessionId": session_id,
                },
                {
                    "upn": upn,
                    "content": [{"type": "text", "text": bilingual_text}],
                    "role": "assistant",
                    "submode": "web",
                    "model": "gpt-4o-transcribe",
                    "from": sendFrom,
                    "sessionId": session_id,
                },
            ]
        }

        # https://itc-history-functions.azurewebsites.net
        url = f"{history_base_url}/api/history/minutes"

        api_name = "add_history"
        try:
            response = await async_http_client.post(
                url=url,
                api_key=history_api_key,
                json_data=request_data,
                process_name=api_name,
            )
        except Exception as e:
            logging.warning(f"履歴の追加: {e}")
        else:
            json_response["session_id"] = response["sessionId"]
            title = response.get("title", "タイトルの取得に失敗しました。")

        json_response["title"] = title
    return func.HttpResponse(json.dumps(json_response))


######
# OCR
######
### input
# data = {
#     "upn": <upn>,
#     "inputs": {
#         "content": ["<Base64エンコードされたデータ>"],
#         "file_name": ["<ファイル名(sample.pdf)>"],
#     }
# }
@app.route(route="genie/ocr", methods=("POST",))
async def ocr(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("OCR processed a request.")
    ########
    # init #
    ########
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    # apiの呼び出し
    history_base_url = os.environ.get("HISTORY_API_URL")
    history_api_key = os.environ.get("HISTORY_API_KEY")

    # get client ip
    client_ip = req.headers.get("x-forwarded-for")
    logging.info(f"client_ip: {client_ip}")

    # get json input
    req_json = req.get_json()
    json_response = {"status": 500}
    upn = req_json["upn"]
    logging.info("upn: " + upn)

    session_id = req_json.get("session_id")
    logging.info(f"session_id: {session_id}")

    try:
        mail = req_json["mail"]
    except Exception as e:
        logging.warning(f"no mail: {e}")

    try:
        sendFrom = req_json["from"]
        assert sendFrom in ["web", "teams", "mail", "agent"]
    except Exception as e:
        req_json["from"] = sendFrom = "web"

    encoded_data = req_json["inputs"]["content"][0]

    # file_name
    file_name = req_json["inputs"]["file_name"][0]
    logging.debug(f"uploaded: {file_name}")

    blobs = None

    ########
    # main #
    ########
    try:
        messages = [
            {
                "role": "user",
                "content": [{"type": "file", "name": file_name, "data": encoded_data}],
            }
        ]
        formatted_messages, blobs = await messages2textMessages(
            upn, messages, async_http_client, image_required=False
        )
        ocr_content = next(
            (
                content["text"]
                for content in formatted_messages[-1]["content"]
                if content["type"] == "text"
            ),
            "",
        )

        json_response["text"] = ocr_content
        json_response["page"] = blobs[0]["page"]
        json_response["status"] = 200
    except Exception as e:
        logging.critical(f"Error in DI: {e}")
        json_response["error"] = f"DI ERROR, {e}"

    if session_id is not None and blobs is not None:
        request_data = {
            "items": [
                {
                    "sessionId": session_id,
                    "upn": upn,
                    "content": [
                        {
                            "type": "blob",
                            "name": blobs[0]["name"],
                            "file_name": blobs[0]["file_name"],
                        }
                    ],
                    "role": "user",
                    "submode": "",
                    "model": "di-ver3.1",
                    "from": sendFrom,
                }
            ]
        }

        # history_base_url: https://itc-history-functions.azurewebsites.net
        url = f"{history_base_url}/api/history/ocr"

        api_name = "add_history"
        try:
            response = await async_http_client.post(
                url=url,
                api_key=history_api_key,
                json_data=request_data,
                process_name=api_name,
            )
        except Exception as e:
            logging.warning(f"履歴の追加: {e}")
        else:
            json_response["session_id"] = response["sessionId"]

    return func.HttpResponse(json.dumps(json_response))


# ocr_prompt
@app.route(route="genie/ocr/prompt", methods=("GET", "POST"))
async def ocr_prompt(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("ocr_prompt processed a request.")
    req_body = req.get_json()
    ocr_mode = req_body.get("ocr_mode", "ocr")
    if ocr_mode == "business_ocr":  # 業務パターン用OCRの場合
        return func.HttpResponse(json.dumps(ocr_prompt_list_business_pattern))
    elif ocr_mode == "shipping_ocr":
        return func.HttpResponse(json.dumps(ocr_prompt_list_shipping))
    else:
        return func.HttpResponse(json.dumps(ocr_prompt_list))


# ocr_response
@app.route(route="genie/ocr/response", methods=("POST",))
async def ocr_response(req: func.HttpRequest) -> func.HttpResponse:
    # data = {
    #     "upn": "",
    #     "mail":"",
    #     "prompt_id": int(<選択されたプロンプトのid>),
    #     "prompt": "<選択されたプロンプト>",
    #     "ocr": "<ocrの内容>"
    # }
    logging.info("OCR_RESPONSE processed a request.")
    ########
    # init #
    ########
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    # apiの呼び出し
    history_base_url = os.environ.get("HISTORY_API_URL")
    history_api_key = os.environ.get("HISTORY_API_KEY")

    # get json input
    req_json = req.get_json()
    try:
        prompt_id = req_json["prompt_id"]
    except:
        logging.warning("no prompt_id")
        prompt_id = 0
    logging.info(f"ocr_prompt_id: {prompt_id}")

    system_content = req_json["prompt"]
    try:
        ocr_content = req_json["ocr"]
    except KeyError:
        pass
    # logging.info("User input:" + user_input)

    upn = req_json["upn"]
    session_id = req_json.get("session_id")
    logging.info(f"session_id: {session_id}")

    try:
        sendFrom = req_json["from"]
        assert sendFrom in ["web", "teams", "mail", "agent"]
    except Exception as e:
        req_json["from"] = sendFrom = "web"

    if session_id:
        params = {"sessionId": session_id}
        url = f"{history_base_url}/api/history/ocr/{upn}"

        api_name = "get_history"
        try:
            ocr_history_json = await async_http_client.get(
                url=url, api_key=history_api_key, params=params, process_name=api_name
            )
        except Exception as e:
            tb = traceback.format_exc()
            error = f"GET_OCR_HISTORY: {e}, {tb}"
            logging.critical(error)
            ocr_response = error_response("処理中にエラーが発生しました。")
            return func.HttpResponse(json.dumps(ocr_response))
        else:
            ocr_history_messages = [
                {"role": "user", "content": ocr_history_json["messages"][0]["content"]}
            ]
            formatted_messages, _ = await messages2textMessages(
                upn, ocr_history_messages, async_http_client
            )
            ocr_content = next(
                (
                    content["text"]
                    for content in formatted_messages[-1]["content"]
                    if content["type"] == "text"
                ),
                "",
            )

    # aoai 関連の初期化
    model_name = "gpt4.1"
    csv_model_name = "gpt4.1-mini"
    _max_tokens = NON_CHAT_REGISTRY.models[model_name].max_tokens
    max_tokens = _max_tokens["input"] - _max_tokens["output"]

    num_tokens = check_token(system_content + ocr_content)
    logging.info(f"OCR_TOKENS: {num_tokens}")
    if max_tokens < num_tokens:
        return func.HttpResponse(
            json.dumps(error_response("ocrの文字数が多すぎます。"))
        )

    ########
    # main #
    ########
    messages = [
        {"role": "system", "content": system_content},
        {"role": "user", "content": ocr_content},
    ]
    try:
        ocr_response = await AzureOpenAI(
            messages=messages,
            temperature=0,
            model_name=model_name,
            max_retries=2,
            timeout=230,
            raise_for_error=False,
            registry=NON_CHAT_REGISTRY,
        )
    except Exception as e:
        logging.critical(f"OCR_RESPONSE: {e}")
        ocr_response = error_response(
            "サーバーからの応答がありません。時間をおいてお試しください。"
        )
        ocr_response["csv"] = ""
        ocr_response["xlsx"] = ""
        return func.HttpResponse(json.dumps(ocr_response))

        # 改行がない場合、改行を追加する
    if "\n" not in ocr_response["choices"][0]["message"]["content"]:
        ocr_response["choices"][0]["message"]["content"] = insert_newline_corrected(
            ocr_response["choices"][0]["message"]["content"]
        )

    # csv
    try:
        response_text = ocr_response["choices"][0]["message"]["content"]
        ocr_csv_system_content = OCR_CSV_SYSTEM_CONTENT
        ocr_csv_system_content = ocr_csv_system_content.format(response=response_text)
        messages = [
            {"role": "system", "content": ocr_csv_system_content},
        ]
        ocr_csv_response = await AzureOpenAI(
            messages=messages,
            temperature=0,
            model_name=csv_model_name,
            max_retries=2,
            timeout=230,
            raise_for_error=False,
            registry=NON_CHAT_REGISTRY,
        )
        # CSVの部分のみ抽出
        ocr_response["csv"] = (
            ocr_csv_response["choices"][0]["message"]["content"]
            .replace("，", ",")
            .replace("```csv", "```")
            .replace("```\n", "```")
            .replace("\n```", "```")
            .split("```")[-2]
        )
    except Exception as e:
        logging.critical(f"CSV ERROR: {e}")
        ocr_response["csv"] = ""

    try:
        # Excelファイルの作成
        wb = Workbook()

        # 1シート目: OCR結果
        ws1 = wb.active
        ws1.title = "OCR結果"

        for line in ocr_content.replace("<br>", "\n").split("\n"):
            # テーブルのマークダウン記法を検出
            if line.startswith("|") and line.endswith("|"):
                # 行が '|', '-', ':'のみで構成されている場合はパス
                if (
                    set(line.replace("|", "").replace("-", "").replace(":", "").strip())
                    == set()
                ):
                    continue
                # 先頭と末尾の'|'を削除し、'|'で分割
                cells = line.strip("|").split("|")
                # 各要素をトリムしてセルに追加
                ws1.append([cell.strip() for cell in cells])
            else:
                # 通常の行として追加
                ws1.append([line])

        # 2シート目: LLMによる変換結果
        ws2 = wb.create_sheet("プロンプトによる変換結果")
        for line in response_text.split("\n"):
            ws2.append([line])

        # 3シート目: LLMによるCSV出力結果
        ws3 = wb.create_sheet("プロンプトによる変換結果(表形式)")

        with io.StringIO(ocr_response["csv"]) as csv_data:
            csv_reader = csv.reader(csv_data)
            for row in csv_reader:
                if len(row) >= 1:
                    # 不要な引用符を削除
                    row = [cell.strip().strip('"').strip("'") for cell in row]
                    # 整数の場合はintにキャスト
                    try:
                        row = [int(cell) if cell.isdigit() else cell for cell in row]
                    except Exception as e:
                        logging.warning(f"csv isdigit failure: {e}")
                    ws3.append(row)

        with io.BytesIO() as excel_buffer:
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            excel_base64 = base64.b64encode(excel_buffer.read()).decode("utf-8")

            # レスポンスにxlsxフィールドを追加
            ocr_response["xlsx"] = excel_base64
    except Exception as e:
        logging.critical(f"EXCEL_GENERATION: {e}")
        ocr_response["xlsx"] = ""

    if session_id:
        request_data = {
            "items": [
                {
                    "sessionId": session_id,
                    "upn": upn,
                    "content": [
                        {"type": "text", "text": system_content},
                        {
                            "type": "blob",
                            "name": ocr_history_messages[0]["content"][0]["name"],
                        },
                    ],
                    "role": "user",
                    "submode": "",
                    "model": model_name,
                    "from": sendFrom,
                },
                {
                    "sessionId": session_id,
                    "upn": upn,
                    "content": [
                        {
                            "type": "text",
                            "text": ocr_response["choices"][0]["message"]["content"],
                        }
                    ],
                    "role": "assistant",
                    "submode": "",
                    "model": model_name,
                    "from": sendFrom,
                },
            ]
        }

        # history_base_url: https://itc-history-functions.azurewebsites.net
        url = f"{history_base_url}/api/history/ocr"

        api_name = "add_history"
        try:
            response = await async_http_client.post(
                url=url,
                api_key=history_api_key,
                json_data=request_data,
                process_name=api_name,
            )
        except Exception as e:
            logging.warning(f"履歴の追加: {e}")
        else:
            ocr_response["session_id"] = response["sessionId"]

    return func.HttpResponse(json.dumps(ocr_response))


# prompt
@app.route(route="genie/prompt", methods=("GET", "POST"))
async def prompt_(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("prompt processed a request.")
    # GETリクエストの場合の処理
    if req.method == "GET":
        query = req.params.get("query", "")

        # language
        try:
            lang = req.params.get("lang")
            assert lang in ["ja", "en"]
        except Exception as e:
            logging.warning(f"lang : {e}")
            lang = "ja"
    elif req.method == "POST":
        try:
            req_json = req.get_json()
            query = req_json["query"]
        except:
            logging.debug("no query")
            query = ""

        # language
        try:
            lang = req_json["lang"]
            assert lang in ["ja", "en"]
        except Exception as e:
            logging.warning(f"lang : {e}")
            lang = "ja"
    # query
    response = query_prompt(query, lang)
    return func.HttpResponse(json.dumps(response))


@app.route(route="genie/prompt/v2", methods=("GET", "POST"))
# @app.route(route="genie/prompt/v2", methods=("GET", "POST"))
async def prompt_v2(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("prompt v2 processed a request.")
    async_http_client = AsyncHttpClient()
    upn = req.params.get("upn")
    id_token = req.params.get("id_token")
    if not id_token:
        return func.HttpResponse(
            json.dumps({"status": 400, "error": "id_token is required"}),
            status_code=400,
        )

    # id_tokenから認証情報を取得
    try:
        keys = await EntraIDTokenManager.get_entra_openid_keys(async_http_client)
        upn_from_token, mail, _ = decode_id_token(id_token, keys)
    except Exception as e:
        logging.warning(f"id_token処理エラー: {e}")

    # GETリクエストの場合の処理
    if req.method == "GET":
        query = req.params.get("query", "")

        # language
        try:
            lang = req.params.get("lang")
            assert lang in ["ja", "en"]
        except Exception as e:
            logging.warning(f"lang : {e}")
            lang = "ja"
    elif req.method == "POST":
        try:
            req_json = req.get_json()
            if upn == None:
                upn = req_json.get("upn")
            query = req_json["query"]
        except:
            logging.debug("no query")
            query = ""

        # language
        try:
            lang = req_json["lang"]
            assert lang in ["ja", "en"]
        except Exception as e:
            logging.warning(f"lang : {e}")
            lang = "ja"

    # id_tokenの検証が成功した場合は検証済みUPNを使用
    if "upn_from_token" in locals():
        upn = upn_from_token

    logging.info(f"prompt_v2 upn: {upn}")

    # お気に入り
    favorite_list = []
    if upn != None:
        prompt_manager = FavoritePromptManager(upn)
        if req.method == "GET":
            favorite_list = prompt_manager.get_favorite_list()
        elif req.method == "POST":
            prompt_id = req_json.get("prompt_id")
            favorite = req_json.get("favorite", 0)
            if prompt_id != None:
                prompt_id = str(prompt_id)
                if bool(favorite):
                    logging.info(f"favorite: add {prompt_id}")
                    prompt_manager.add_favorite(prompt_id)
                else:
                    logging.info(f"favorite: del {prompt_id}")
                    prompt_manager.remove_favorite(prompt_id)
            favorite_list = prompt_manager.get_favorite_list()

    # query
    try:
        response = {"status": 200, "data": process_prompt(query, favorite_list)}
    except Exception as e:
        tb = traceback.format_exc()
        logging.critical(f"error in prompt_v2: {e}, tb: {tb}")
        response = {"status": 500, "error": f"{e}"}
    return func.HttpResponse(json.dumps(response))


# test
@app.route(route="sendmail", methods=("GET", "POST"))
async def sendmail(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("sendmail processed a request.")
    # upn = "A227002"

    if req.method == "GET":
        try:
            mail = req.params.get("mail")
            if "@" in mail:
                mail = mail.split("@")[0]
            mail += "@itochu.co.jp"
        except:
            mail = "takeuchi-yuki@itochu.co.jp"

        subject = ("【I-Colleague】メール送信テストです。",)
        content = "開発環境のGenieからメールを送信しています。\nこちらで「生成AIラボ」という名前が表示されるように設定を追加しております。\nよろしくお願いいたします。"

        attachment_list = [
            {
                "file_name": "minutes.txt",
                "file_content": "議事録です。".encode("utf-8"),
            },
            {
                "file_name": "transcribe.txt",
                "file_content": "文字起こしです。".encode("utf-8"),
            },
        ]

    elif req.method == "POST":
        req_json = req.get_json()

        # get params
        mail = req_json["mail"]
        if "@" in mail:
            mail = mail.split("@")[0]
        mail += "@itochu.co.jp"

        subject = "【I-Colleague】" + req_json["subject"]
        content = req_json["content"]

        attachment_list = []
        if "attachment_list" in req_json.keys():
            attachment_list = req_json["attachment_list"]

            # エンコードされたデータの場合を受け取った場合はバイナリデータに変換
            for item in attachment_list:
                if item.get("encode_type") == "base64":
                    item["file_content"] = base64.b64decode(item["file_content"])

    try:
        flag = send_email(
            subject=subject,
            content=content,
            send_to=[
                mail,
            ],
            attachment_list=attachment_list,
        )
    except Exception as e:
        logging.info("メールの送信に失敗しました:" + str(e))
        flag = "メールの送信に失敗しました:" + str(e)

    return func.HttpResponse(f"{flag}, mail: {mail}")


@app.route(route="durable/{functionName}")
@app.durable_client_input(client_name="client")
async def start_orchestrator(req: func.HttpRequest, client):
    """
    durable functionsの呼び出し用のAPI
    postされたデータを読み込み、urlにて指定された関数を呼び出す
    """

    payload: dict = json.loads(req.get_body().decode())  # Load JSON post request data
    instance_id = await client.start_new(
        req.route_params["functionName"], client_input=payload
    )

    logging.info(f"Started orchestration with ID = '{instance_id}'.")
    return client.create_check_status_response(req, instance_id)


@app.route(route="genie/merchant", methods=("GET",))
async def send_merchant_rate(req: func.HttpRequest) -> func.HttpResponse:
    async def manage_merchant_rate():
        """
        次回更新の時間を過ぎている場合は最新のものを取得
        """
        global MERCHANT_RATE
        # 現在の時刻を UTC で取得
        next_update_str = MERCHANT_RATE.get("next_update", "0001-01-01T00:00:00")
        next_update = datetime.fromisoformat(next_update_str)

        # タイムゾーン情報の確認
        if next_update.tzinfo is None:
            next_update = next_update.replace(tzinfo=timezone.utc)

        # 現在の時刻を UTC のタイムゾーン情報付きで取得
        current_time = datetime.now(timezone.utc)

        if current_time > next_update:
            # 更新処理
            MERCHANT_RATE = await access_merchant_rate("get")

        return MERCHANT_RATE.get("daily_user", 0)

    logging.info("genie top processed a request.")
    global MERCHANT_RATE
    try:
        lang = req.params.get("lang", "ja")

        response = {
            "status": 200,
            "data": {
                "merchant": await manage_merchant_rate(),
            },
        }
    except Exception as e:
        tb = traceback.format_exc()
        error = f"merchant error: {e}, {tb}"
        logging.critical(error)

        response = {"status": 500, "error": error}
    return func.HttpResponse(json.dumps(response))


@app.route(route="genie/top", methods=("GET",))
async def send_top_page_content(req: func.HttpRequest) -> func.HttpResponse:
    """
    status:
    - 200: 成功
    - 500: 失敗

    新UIのトップページの情報を受け渡す
    - 今月/今日の商人率: merchant
    - おすすめのプロンプト: prompts
    - 新着情報: news

    ```
    {
        "status": 200,
        "data": {
            "merchant": 100,
            "messages": [
                    {
                        "category": "dog",
                        "content": "こんにちは"
                    },
                    {
                        "category": "welcome",
                        "content": "ようこそ！生成AIの世界へ。"
                    }
                ],
            "prompts": [
                {
                    "category": "事務作業",
                    "title": "Excelでやりたいことを実現する方法を教えてもらう",
                    "prompt":"# 命令書 ...",
                    "tag": ["説明・教えて", ]
                },
                {},
                {},
                {}
            ],
            "news": [
                {
                    "date": "9/20",
                    "category": "アップデートのお知らせ",
                    "title": "I-Colleagueのリニューアル！",
                    "content": "I-ColleagueはX月X日に...",
                    "link": "https://example.com"
                },
                {},
                {},
                {}
            ]
        }
    }
    ```

    """

    def choice_dog_message():
        if lang == "ja":
            # 通常時のメッセージ
            now_jst = datetime.now(ZoneInfo("Asia/Tokyo"))
            message = 'お疲れ様です<span class="color-key">♪</span>'
            if 0 <= now_jst.hour < 5:
                message = 'お疲れ様です<span class="color-key">♪</span>'
            elif 5 <= now_jst.hour < 10:
                message = 'おはようございます<span class="color-key">！</span>'
            elif 10 <= now_jst.hour < 17:
                message = 'こんにちは<span class="color-key">！</span>'
            elif 17 <= now_jst.hour < 24:
                message = 'お疲れ様です<span class="color-key">♪</span>'

            # 期間限定のメッセージ
            if now_jst.month == 1 and now_jst.day <= 10:
                message = random.choice(
                    [
                        '謹賀新年<span class="color-key">！</span>',
                        'あけおめワン<span class="color-key">！</span>',
                    ]
                )
            elif now_jst.month == 10 and now_jst.day == 31:
                message = 'Happy Halloween<span class="color-key">！</span>'
            elif now_jst.month == 12 and now_jst.day in (24, 25):
                message = 'メリークリスマス<span class="color-key">！</span>'
            elif now_jst.month == 12 and now_jst.day >= 26:
                message = 'よいお年を<span class="color-key">！</span>'
        else:
            message = 'Hello<span class="color-key">♪</span>'
        return message

    def choice_welcome_message():
        if lang == "ja":
            messages = [
                "さっきの会議でわからなかったこと、こっそり私に聞いてください",
                "それ、一旦私に聞いてみませんか？",
                "何かお困りですか？",
                "何でもお手伝いします！",
                "今日は早めにあがりましょう、何でも手伝います！",
                "要件を聞かせて！",
                "ようこそ！生成AIの世界へ。",
                "今日も頑張りましょう！",
                "その資料、私が要約／翻訳します！",
                "生成AI、始めるならI-Colleague。",
                "あなたの相棒、I-Colleagueです！",
                "どんなことでもお手伝いします。今日は何をしますか？",
                "「まず、I-Colleague。」",
                "おかえりなさい、今日も頑張りましょう！",
            ]
        else:
            messages = [
                "These kinds of prompts are recommended!",
                "If there's anything you didn't understand from the earlier meeting, feel free to ask me",
                "Why don’t you ask me what you have in mind?",
                "Is there anything I can help you with?",
                "I will help with anything!",
                "Let’s finish up early today; I’ll help with anything!",
                "What can I do for you?",
                "Welcome to the world of generative AI!",
                "Let’s give it our all today!",
                "I’ll summarize/translate your document!",
                "If you’re starting with generative AI, go with I-Colleague",
                "Your partner, I-Colleague!",
                "I will help with anything. What’s on the agenda for today?",
                "First, I-Colleague",
                "Welcome back! Let’s do our best today!",
            ]
        return random.choice(messages)

    logging.info("genie top processed a request.")
    global MERCHANT_RATE
    try:
        lang = req.params.get("lang", "ja")

        response = {
            "status": 200,
            "data": {
                "merchant": 100,
                "messages": [
                    {"category": "dog", "content": choice_dog_message()},
                    {"category": "welcome", "content": choice_welcome_message()},
                ],
                "prompts": choice_prompt(),
                "news": get_news(),
            },
        }
    except Exception as e:
        tb = traceback.format_exc()
        error = f"top error: {e}, {tb}"
        logging.critical(error)

        response = {"status": 500, "error": error}
    return func.HttpResponse(json.dumps(response))


@app.route(route="genie/auth/{mode}", methods=("GET",))
async def check_user_authority(req: func.HttpRequest) -> func.HttpResponse:
    """
    ユーザーの所属情報かメニュー表示の権限を取得して渡す。

    Request Body:
        - routeパラメータ
            - mode (str): "db"または"menu"。
        - クエリパラメータ
            - upn (str): ユーザーのユーザープリンシパルネーム

    Returns:
        - func.HttpResponse: 以下のいずれかの情報を含むHTTP応答オブジェクト。
            - modeが"db"の場合: ユーザーの属性情報（user_attributes）

            ```
            {
                "status": 200,
                "data": {
                    "attributes": [ "A227003@intra.itochu.co.jp", "IT・デジタル戦略部", "CXO" ]
                }
            }
            ```
            - modeが"menu"の場合: ユーザーの権限に基づくメニューの許可情報

            ```
            {
                "status": 200,
                "data": {
                    "permissions": {
                        "allowed": ["top", "default_gpt", ...],
                        "guide": ["prompt", ...]
                    }
                }
            }
            ```
            - 例外発生時: エラー情報を含む応答

            ```
            {
                "status": 500,
                "error": "<エラーメッセージ>"
            }
            ```
    """
    mode = req.route_params.get("mode")
    logging.info(f"check auth: {mode}")
    upn = req.params.get("upn")
    logging.info(f"upn: {upn}")

    # 所属情報の取得
    client = UserDivisionFetchService()
    user_attributes = client.fetch_user_attributes(upn)
    logging.info(f"user_attribute: {user_attributes}")

    response = {"status": 500, "error": f"invalid mode: {mode}"}

    if mode == "db":
        response = {
            "status": 200,
            "data": {"attributes": user_attributes},
        }

    elif mode == "menu":
        try:
            # メニュー表示権限の取得
            menu_client = MenuPermissionService(cosmos_client=COSMOS_CLIENT)
            permissions = menu_client.fetch_menu_permissions(user_attributes)
            logging.info(f"permissions: {permissions}")

            response = {
                "status": 200,
                "data": {"permissions": permissions},
            }

        except Exception as e:
            tb = traceback.format_exc()
            error = f"error: {e}, tb: {tb}"
            logging.critical(f"Auth {error}")
            response = {"status": 500, "error": error}

    return func.HttpResponse(json.dumps(response))


# id_tokenのデコード
@app.route(route="decode", methods=("POST",))
async def decode_idtoken(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("decode_idtoken: Start.")
    # 非同期http通信クライアントのインスタンス化
    async_http_client = AsyncHttpClient()

    # ip情報の取得
    client_ip = req.headers.get("x-forwarded-for")
    logging.info(f"client_ip: {client_ip}")
    # id_token取得
    req_json = req.get_json()
    id_token = req.params.get("id_token")

    # id_tokenのデコード
    if id_token:
        try:
            keys = await EntraIDTokenManager.get_entra_openid_keys(async_http_client)
            upn, mail, _ = decode_id_token(id_token, keys)
            logging.info("id_tokenのデコードに成功")
        except Exception as e:
            logging.warning("id_tokenのデコードに失敗: {e}")
            response = {"status": 200, "error": "id_tokenのデコードに失敗"}
            return func.HttpResponse(json.dumps(response), status_code=200)
    else:
        logging.warning("id_token not found.")
        response = {"status": 200, "error": "id_tokenが見つかりません"}
        return func.HttpResponse(json.dumps(response), status_code=200)

    # デコードしたものでreq_jsonを上書き
    req_json["upn"] = upn
    req_json["mail"] = mail

    mode = req_json.get("mode")

    logging.info(f"upn: {upn}")
    logging.info(f"mode: {mode}")

    # DurableFunctionへリクエスト
    try:
        status_code, durable_response = await async_http_client.post(
            url=MCP_AGENT_URL,
            api_key=MCP_AGENT_API_KEY,
            json_data=req_json,
            return_status=True,
        )
        return func.HttpResponse(
            json.dumps(durable_response),
            status_code=status_code,
            mimetype="application/json",
        )
    except Exception as e:
        logging.warning(f"Durable Functionとの通信に失敗しました: {e}")
        response = {
            "status": 200,
            "error": "サーバーからの応答がありません。時間をおいてお試しください。",
        }
        return func.HttpResponse(json.dumps(response), status_code=200)
