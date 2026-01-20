import azure.functions as func
import azure.durable_functions as d_func
from difflib import HtmlDiff
from bs4 import BeautifulSoup, Tag, NavigableString
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock
from itertools import zip_longest
import logging
import base64
import re
import json
import io

##
# blueprint
##
file_diff_bp = d_func.Blueprint()


class DiffProcessor:
    """2つのテキストの差分を解析するクラス。"""

    def __init__(self, file1_splitted, file2_splitted, file1_name, file2_name):
        """
        Args:
            file1 (list[str]): file1を句点、改行コードで分割したもの。
            file2 (list[str]): file2を句点、改行コードで分割したもの。
            file1_name (str): file1のファイル名。
            file2_name (str): file2のファイル名。
        """
        self.file1 = file1_splitted
        self.file2 = file2_splitted
        self.file1_name = file1_name
        self.file2_name = file2_name
        self.html = self._generate_html_diff()
        self.soup = BeautifulSoup(self.html, "lxml")

    def _generate_html_diff(self):
        """difflibを使って差分HTMLを生成する。"""
        hd = HtmlDiff(tabsize=4, wrapcolumn=0)
        return hd.make_file(
            self.file1, self.file2, self.file1_name, self.file2_name,
            context=False, numlines=5
        )

    def iter_diff_rows(self):
        """差分HTMLテーブルの各行を解析し、左右のセルデータと差分要約を生成するジェネレータ。"""
        table = self.soup.find("table")
        for tr in table.find_all("tr"):
            nowrap_cells = tr.find_all("td", attrs={"nowrap": True})
            if not nowrap_cells:
                continue

            left_td, right_td = nowrap_cells[0], nowrap_cells[1]
            runs_l = self._extract_runs(left_td)
            runs_r = self._extract_runs(right_td)
            summary = self._build_summary(runs_l, runs_r)

            yield {
                "runs_left": runs_l,
                "runs_right": runs_r,
                "summary": summary
            }

    def _extract_runs(self, td_element):
        """<td>要素からテキストと差分タグのペア [(tag, text), ...] を抽出する。"""
        runs = []
        for node in td_element.contents:
            if isinstance(node, NavigableString):
                runs.append(("normal", str(node)))
            elif isinstance(node, Tag) and node.get("class"):
                cls = node["class"][0]
                txt = node.get_text()
                if cls == "diff_sub":
                    runs.append(("sub", txt))
                elif cls == "diff_add":
                    runs.append(("add", txt))
                elif cls == "diff_chg":
                    runs.append(("chg", txt))
                else:
                    runs.append(("normal", txt))
            else:
                # その他のタグが存在した場合はnormalとして扱い、中身だけ抽出
                runs.append(("normal", node.get_text()))
        return runs

    def _build_summary(self, runs_left, runs_right):
        """差分から相違点(削除、追加、変更)のまとめを作成する。"""
        subs = [txt for tp, txt in runs_left if tp == "sub"]
        adds = [txt for tp, txt in runs_right if tp == "add"]
        chg_pairs = list(zip_longest(
            [txt for tp, txt in runs_left if tp == "chg"],
            [txt for tp, txt in runs_right if tp == "chg"],
            fillvalue=""
        ))

        parts = []
        if subs:
            parts.append(f"削除:{','.join(f'{s}' for s in subs)}")
        if adds:
            parts.append(f"追加:{','.join(f'{a}' for a in adds)}")
        if chg_pairs:
            parts.append(
                f"変更:{','.join(f'{old}→{new}' for old, new in chg_pairs)}")

        return "\n".join(parts)

    @staticmethod
    def split_sentences(text):
        """
        句点と改行で文を分割し、各文を1行として扱う。

        Args:
            text(str): 文字起こし結果。
        Returns:
            list[str]: 分割された文のリスト。入力がstr出ない場合は空のリストを返却する。
        """
        if not isinstance(text, str):
            return []

        s = text.replace('\r\n', '\n').replace('\r', '\n')
        pattern = re.compile(r'[^。\n]+(?:。)?')
        return [m.group(0).strip() for m in pattern.finditer(s) if m.group(0).strip()]


class ExcelService():
    """差分解析結果をExcelファイルに書き込むクラス。"""

    DIFF_SHEET_NAME = "比較結果"
    DIFF_HEADER_COL3 = "相違点"
    FONT_COLORS = {
        "normal": "000000",
        "add": "006100",
        "sub": "9C0006",
        "chg": "CC7A00",
    }

    def __init__(self):
        self.wb = self._create_workbook()
        self._inline_fonts = {
            tag: InlineFont(color=color) for tag, color in self.FONT_COLORS.items()
        }

    def _create_workbook(self):
        """新しいワークブックを作成する。"""
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def _validate_sheet_title(self, title):
        """Excelシート名として無効な文字を置換し、長さを調整する。"""
        # 使用できない文字を指定の文字（例: "_"）に置換
        sanitized_title = re.sub(r'[\\/*?:\[\]]', '_', title)

        # 先頭と末尾のシングルクォートを削除
        sanitized_title = sanitized_title.strip("'")

        # シート名の長さを31文字以内(Excelの仕様)に制限
        if len(sanitized_title) > 31:
            sanitized_title = sanitized_title[:31]
        return sanitized_title

    def resolve_sheet_names(self, name1, name2):
        """2つのファイル名が衝突しないか確認する。衝突する場合は、衝突しないExcelシート名を生成する。"""
        validated_name1 = ExcelService._validate_sheet_title(self, name1)
        validated_name2 = ExcelService._validate_sheet_title(self, name2)

        if validated_name1 == validated_name2:
            return f"{validated_name1[:28]}_01", f"{validated_name2[:28]}_02"
        else:
            return validated_name1, validated_name2

    def add_source_sheet(self, lines, title, sheet_name):
        """元のテキストデータを格納するシートを追加する。"""
        ws = self.wb.create_sheet(title=sheet_name)

        # ファイル名を1行目に格納
        title_cell = ws.cell(row=1, column=1, value=title)
        # 折り返し表示
        title_cell.alignment = Alignment(wrap_text=True, vertical='top')
        for i, line in enumerate(lines, start=2):
            value = (
                f"'{line}" if line.startswith("=") else line  # =で始まる場合は'を付与
            )
            cell = ws.cell(row=i, column=1, value=value)
            # 折り返し表示
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    def add_diff_sheet(self, diff_iterator, file1_name, file2_name):
        """差分比較結果を格納するシートを追加する。"""
        ws = self.wb.create_sheet(title=self.DIFF_SHEET_NAME)
        # 各ファイル名を1行目に格納
        ws.cell(row=1, column=1, value=file1_name)
        ws.cell(row=1, column=2, value=file2_name)
        ws.cell(row=1, column=3, value=self.DIFF_HEADER_COL3)

        for row_idx, row_data in enumerate(diff_iterator, start=2):
            # 比較元格納セル
            self._write_text_cell(ws, row_idx, 1, row_data["runs_left"])
            self._write_text_cell(ws, row_idx, 2, row_data["runs_right"])
            # 相違点格納セル
            diff_cell = ws.cell(row=row_idx, column=3,
                                value=row_data["summary"])
            # 折り返し表示
            diff_cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _write_text_cell(self, ws, row, col, runs):
        """指定されたセルにテキストを書き込む。"""
        cell = ws.cell(row=row, column=col)
        if not runs:
            return

        blocks = [TextBlock(self._inline_fonts[tag], text)
                  for tag, text in runs if text]
        if len(blocks) == 1 and runs[0][0] == "normal":
            cell.value = blocks[0].text
        elif blocks:
            cell.value = CellRichText(blocks)
        # 折り返し表示
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _format(self):
        """全てのシートに対し、書式設定及び列幅を調整する。"""
        for ws in self.wb.worksheets:
            self._set_number_formats(ws)
            self._adjust_column_width(ws)

    def _set_number_formats(self, ws):
        """セルの内容に応じて数値書式または文字列書式を設定する。"""
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "General"
                else:
                    cell.number_format = "@"

    def _adjust_column_width(self, ws):
        """ワークシートの列幅を内容に合わせて自動調整する。"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            # 列の全セルをチェック
            for cell in column:
                try:
                    # セルの値を文字列として取得
                    cell_value = str(
                        cell.value) if cell.value is not None else ""
                    # 現在の最大長と比較
                    max_length = max(max_length, len(cell_value))
                except Exception as e:
                    logging.warning(
                        f"Error processing cell {cell.coordinate}: {str(e)}")
                    continue
            # 最小幅と最大幅の制限を設定 (最小8文字、最大50文字)
            adjusted_width = min(max(max_length + 2, 8), 50)
            # 列幅を設定
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_bytes(self):
        """ワークブックをメモリ上のバイトデータとして取得する。"""
        self._format()
        buffer = io.BytesIO()
        self.wb.save(buffer)
        return buffer.getvalue()


@file_diff_bp.route(route="genie/file_diff", methods=("POST",))
async def file_diff(req: func.HttpRequest) -> func.HttpResponse:
    """2つのテキストファイルを受け取り、その差分をHTML及びExcelファイルで出力する関数。"""
    logging.info("file_diff: Start.")

    # リクエストの受けとり
    try:
        client_ip = req.headers.get("x-forwarded-for")
        req_json = req.get_json()
        file1 = req_json["file1"]
        file1_name = req_json["file1_name"]
        file2 = req_json["file2"]
        file2_name = req_json["file2_name"]
        upn = req_json["upn"]
    except Exception as e:
        logging.error(f"リクエストの受付処理時エラー: {e}")
        res_json = {"message": "リクエストの受付処理に失敗しました"}
        return func.HttpResponse(
            json.dumps(res_json),
            mimetype="application/json",
            status_code=500
        )

    try:
        logging.info(f"client_ip: {client_ip}")
        logging.info(f"upn: {upn}")
        # 文字起こし結果を句点と改行で分割
        file1_splitted = DiffProcessor.split_sentences(file1)
        file2_splitted = DiffProcessor.split_sentences(file2)

        # html差分抽出
        diff_processor = DiffProcessor(
            file1_splitted, file2_splitted, file1_name, file2_name)
        html = diff_processor.html

        # Excelファイル作成
        excel_service = ExcelService()
        sheet1_name, sheet2_name = excel_service.resolve_sheet_names(
            file1_name, file2_name)
        excel_service.add_source_sheet(file1_splitted, file1_name, sheet1_name)
        excel_service.add_source_sheet(file2_splitted, file2_name, sheet2_name)
        excel_service.add_diff_sheet(
            diff_processor.iter_diff_rows(), file1_name, file2_name)
        excel_bytes = excel_service.get_bytes()

        # html & Excelファイル返却
        html_base64 = base64.b64encode(html.encode('utf-8')).decode('utf-8')
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        res_json = {
            "message": "差分表示用HTMLとExcelファイルの作成に成功しました。",
            "html_diff_base64": html_base64,
            "excel_diff_base64": excel_base64
        }
        return func.HttpResponse(
            json.dumps(res_json, ensure_ascii=False),
            mimetype="application/json",
            status_code=200
        )
    except Exception as e:
        logging.error(
            f"file_diff: 差分生成またはExcel作成中にエラーが発生しました: {e}", exc_info=True)
        return func.HttpResponse(
            json.dumps({"message": "処理中にエラーが発生しました。"}),
            mimetype="application/json",
            status_code=500
        )
