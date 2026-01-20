import azure.functions as func
import azure.durable_functions as d_func
from azure.storage.blob.aio import BlobServiceClient
import traceback
import logging
import json
import os
import io
import re
import asyncio
from datetime import datetime
from zoneinfo import ZoneInfo
import base64
from openpyxl import Workbook, load_workbook
from typing import Union, List, Dict, Any
from abc import ABC, abstractmethod

# 自作
from config import NON_CHAT_REGISTRY, BLOB_SERVICE_CLIENT
from util import check_token
from prompt import coa_comparison_sub_prompt_list, prompt_shipping_doc_classify, ocr_prompt_list_shipping

from i_style.llm import AzureOpenAI

# blob
BLOB_CONNECTION_STRING = os.environ.get("BLOB_CONNECTION_STRING")
FILE_CONTAINER_NAME = "ocr-business-pattern"

# model
model_gpt4o_mini = "gpt4o-mini"
model_gpt4o = "gpt4o"
model_gpt4_1_mini = "gpt4.1-mini"
model_gpt4_1 = "gpt4.1"
model_o1 = "o1"
model_o3_mini = "o3-mini"
model_o4_mini = "o4-mini"

max_tokens = 128_000 - 4096

##
# blueprint
##
ocr_bp = d_func.Blueprint()


@ocr_bp.route(route="genie/ocr/response2", methods=("POST",))
async def ocr_response2(req: func.HttpRequest) -> func.HttpResponse:
    """
    サンプルデータ
    ```
    data = {
        "upn": "",
        "mail":"",
        "prompt": [{""id": "1", title": "L／C読取", "prompt": "<選択されたプロンプト>"}, ...],
        "ocr": [{"fileName": "<ファイル名>", "content": "<OCR結果>"}, ...],
        "ocr_mode": "<ocrのモード>", # "ocr" or "shipping_ocr" or "business_ocr"
        "json_mode": true or false
    }
    ```
    """
    logging.info('OCR_RESPONSE2 processed a request.')

    # ip情報の取得
    client_ip = req.headers.get("x-forwarded-for")
    logging.info(f"client_ip: {client_ip}")

    try:
        # リクエストデータの準備
        req_json = req.get_json()
        upn = req_json["upn"]
        system_contents = req_json["prompt"]
        ocr_contents = req_json["ocr"]
        ocr_mode = req_json["ocr_mode"]
        business_pattern = req_json.get("business_pattern", "ocr_results")

        logging.info(f"upn: {upn}")

        # プロセッサーの選択
        if ocr_mode == "shipping_ocr":
            processor = ShippingDocumentTextExtractor(upn)
        # elif ocr_mode == "comparison_ocr":
        #     processor = DocumentComparisonProcessor(upn)
        else:
            processor = DocumentProcessor(business_pattern, upn)

        # 文書処理の実行
        excel_files = await processor.process(ocr_contents, system_contents)

        return func.HttpResponse(
            json.dumps(excel_files),
            mimetype="application/json",
            status_code=200
        )

    except Exception as e:
        tb = traceback.format_exc()
        error = f"OCR_RESPONSE: {e}, tb: {tb}"
        logging.critical(error)
        return func.HttpResponse(
            json.dumps({"message": "船積書類の読み込みに失敗しました。", "error": error}),
            mimetype="application/json",
            status_code=500
        )


@ocr_bp.route(route="genie/ocr/classify_documents", methods=("POST",))
async def ocr_classify_documents(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('ocr_classify_documents processed a request.')

    try:
        req_json = req.get_json()
        ocr_contents = req_json.get("ocr", [])

        # 書類分類の実行
        processor = ShippingDocumentTypeIdentifier()
        ocr_classify_documents_results = await processor.process(
            ocr_contents=ocr_contents,
            system_prompts=[prompt_shipping_doc_classify[1]]
        )

        return func.HttpResponse(
            json.dumps(ocr_classify_documents_results),
            mimetype="application/json",
            status_code=200
        )

    except Exception as e:
        tb = traceback.format_exc()
        error = f"OCR_CLASSIFY_DOCUMENTS: {e}, tb: {tb}"
        logging.critical(error)
        return func.HttpResponse(
            json.dumps({"message": "船積書類の判別に失敗しました。", "error": error}),
            mimetype="application/json",
            status_code=500
        )


@ocr_bp.route(route="genie/ocr/upload_prompts", methods=("POST",))
async def upload_prompts(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('upload_prompts processed a request.')

    try:
        # JSONデータを取得
        req_body = req.get_json()
        upn = req_body.get('upn')
        mail = req_body.get('mail')

        inputs = req_body['inputs']
        file_names = inputs.get('file_name', [])
        base64_data_list = inputs.get('data', [])

        file_name = file_names[0]
        base64_data = base64_data_list[0]

        file_content = base64.b64decode(base64_data)

        # ファイルをメモリ上で読み込む
        file_stream = io.BytesIO(file_content)

        workbook = load_workbook(file_stream)
        sheet = workbook.active

        # 許可されたシート名のリスト
        ALLOWED_SHEET_NAMES = [
            "PO指示文",
            "LC指示文",
            "BL指示文",
            "INV指示文",
            "PL指示文",
            "COO指示文",
            "COA指示文",
            "COQ指示文"
        ]

        # 指示文シートの存在チェック
        instruction_sheets = [
            sheet_name for sheet_name in workbook.sheetnames
            if sheet_name in ALLOWED_SHEET_NAMES
        ]

        if not instruction_sheets:
            return func.HttpResponse(
                json.dumps({
                    "message": f"指示文シートが見つかりません。\n使用可能なシート名: {', '.join(ALLOWED_SHEET_NAMES)}",
                    "error": "エラー"
                }),
                mimetype="application/json",
                status_code=400
            )

        prompts = []
        prompt_data = []

        # 許可されたシートのみを処理
        for sheet_name in ALLOWED_SHEET_NAMES:
            logging.info(f"Processing sheet: {sheet_name}")
            sheet_content = ""

            if sheet_name in workbook.sheetnames:
                logging.info(f"Sheet {sheet_name} found in workbook")
                sheet = workbook[sheet_name]

                # シートの内容を1つの文字列として取得
                sheet_content = ""

                # データが存在する範囲を取得
                data_rows = sheet.iter_rows(
                    min_row=1,
                    max_row=sheet.max_row,
                    min_col=1,
                    max_col=sheet.max_column,
                    values_only=True
                )

                # 各行のデータを処理
                for row in data_rows:
                    # None以外の値を結合
                    row_content = " ".join(str(cell)
                                           for cell in row if cell is not None)
                    # 空行の場合は改行のみ追加、内容がある場合は内容と改行を追加
                    sheet_content += row_content + "\n"

                # 最後の余分な改行を削除
                sheet_content = sheet_content.rstrip()

            # シートが存在しないか、内容が空の場合
            if not sheet_content:
                logging.info(
                    f"Sheet {sheet_name} is empty or not found, using default prompt")
                sheet_title = sheet_name.replace("指示文", "")
                logging.info(
                    f"Looking for default prompt with title: {sheet_title}")

                default_prompt = next(
                    (item for item in ocr_prompt_list_shipping if item["title"] == sheet_title),
                    None
                )

                if default_prompt:
                    logging.info(f"Adding default prompt for {sheet_title}")
                    prompts.append({
                        "title": default_prompt["title"],
                        "prompt": default_prompt["prompt"]
                    })
                else:
                    logging.info(f"No default prompt found for {sheet_title}")
            else:
                prompts.append({
                    "title": sheet_name.replace("指示文", ""),
                    "prompt": sheet_content
                })

        prompt_data.append({
            "fileName": file_name,
            "prompts": prompts
        })

        return func.HttpResponse(
            json.dumps(prompt_data),
            mimetype="application/json",
            status_code=200
        )

    except Exception as e:
        tb = traceback.format_exc()
        error = f"OCR_UPLOAD_PROMPTS: {e}, tb: {tb}"
        logging.critical(error)
        return func.HttpResponse(
            json.dumps({"message": "指示文のアップロードに失敗しました。", "error": error}),
            mimetype="application/json",
            status_code=500
        )

##################
# インターフェース #
##################


class BaseDocumentProcessor(ABC):
    """
    文書処理のインターフェース。

    このインターフェースは、文書処理のメソッドを定義する。
    """
    @abstractmethod
    async def process(self, ocr_contents: List[Dict], system_prompts: List[Dict]) -> Dict:
        """文書を処理する"""
        pass

    async def generate_response(self, content: str, system_prompt: str, model_name: str, json_mode: bool) -> Dict:
        """AOAIを呼び出し、回答を生成する。"""
        try:
            if check_token(system_prompt + content) > max_tokens:
                return {"error_type": "token_error", "message": "トークン数が制限を超えています。"}

            response = await AzureOpenAI(
                [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": content},
                ],
                model_name=model_name,
                temperature=0,
                timeout=230,
                json_mode=json_mode,
                raise_for_error=True,
                registry=NON_CHAT_REGISTRY
            )

            if not response or 'choices' not in response:
                raise ValueError("Invalid response format from Azure OpenAI")

            result = json.loads(response['choices'][0]['message']['content']
                                ) if json_mode else response['choices'][0]['message']['content']
            return result

        except Exception as e:
            logging.error(
                f"Error in generate_response: {traceback.format_exc()}")
            return {
                "error_type": str(type(e)),
                "message": str(e),
            }

    async def _process_single_prompt(self, content: str, prompt_config: Dict, model_name: str, json_mode: bool) -> Any:
        """単一のコンテンツにに対して単一のプロンプト処理を行う。"""
        response = await self.generate_response(
            content=content,
            system_prompt=prompt_config["prompt"],
            model_name=model_name,
            json_mode=json_mode
        )

        if isinstance(response, Exception):
            logging.error(f"AOAI Error: {str(response)}")
            response = self.create_error_response(
                content["fileName"],
                f"AOAI Processing error: {str(response)}"
            )

        return self._normalize_data(response)

    async def upload_blob(self, file_name: str, data: bytes) -> None:
        """ファイルをAzure Blob Storageにアップロードする。"""
        try:
            container_client = BLOB_SERVICE_CLIENT.get_container_client(
                FILE_CONTAINER_NAME)

            try:
                await container_client.create_container()
            except Exception:
                pass  # コンテナが既に存在する場合は無視

            blob_client = container_client.get_blob_client(file_name)
            await blob_client.upload_blob(data, overwrite=True)

        except Exception as e:
            logging.error(f"Error in upload: {traceback.format_exc()}")
            raise

    def create_error_response(self, file_name, error_message):
        """エラーレスポンスを生成する。"""
        return {
            "fileName": file_name,
            "error": str(error_message),
            "status": "error"
        }

    def _normalize_data(self, data: Union[List[Dict[str, Any]], Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        データの正規化を行う。
        """
        # 対象となるキーのリスト
        TARGET_KEYS = ['data', 'result', '結果', 'データ']

        # データが空の場合
        if not data:
            return data

        # リストの場合の処理
        if isinstance(data, list):
            # リストの要素が1つの場合
            if len(data) == 1:
                # リストの最初の要素が辞書で、対象キーのいずれかを持つ場合
                if isinstance(data[0], dict):
                    for key in TARGET_KEYS:
                        if key in data[0]:
                            return data[0][key]
                return self._normalize_data(data[0])
            return data

         # 辞書の場合の処理
        if isinstance(data, dict):
            # 対象キーのいずれかが存在する場合、最初に見つかったキーの値を返す
            return next(
                (data[key] for key in TARGET_KEYS if key in data),
                data
            )

        return data

    def _split_content_by_pages(self, content: str) -> List[str]:
        """OCRコンテンツをページ単位に分割する。"""
        file_name_match = re.match(r'##\s*\[.*?\](?:\[.*?\])*', content)
        file_name = file_name_match.group(0) if file_name_match else ""

        content_without_filename = content.replace(file_name, "").strip()
        pages = content_without_filename.split("### Page")

        pages = [page.strip() for page in pages if page.strip()]
        pages = [pages[0]] + ["### Page" +
                              page for page in pages[1:]] if pages else []

        if file_name:
            pages.insert(0, file_name)

        return pages

    def _get_prompt(self, system_prompts: List[Dict[str, Any]], title: str) -> Any:
        """複数のプロンプトリストの中から指定したプロンプトを取得する"""
        for prompt_config in system_prompts:
            if prompt_config["title"] == title:
                return prompt_config
        return None

################
# サービスクラス #
################


class ExcelService():
    """Excel生成サービスクラス"""

    def create_workbook(self) -> Workbook:
        """新しいワークブックを作成する。"""
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def create_unique_sheet(self, wb: Workbook, base_name: str) -> tuple[Workbook, str]:
        """ユニークなシート名でシートを作成する。"""
        sheet_name = self._get_unique_sheet_name(wb, base_name)
        wb.create_sheet(title=sheet_name)
        return wb, sheet_name

    def _get_unique_sheet_name(self, wb: Workbook, base_name: str) -> str:
        """ユニークなシート名を生成する。"""
        sheet_name = base_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name[:27]}_{counter}"
            counter += 1
        return self._validate_sheet_title(sheet_name)

    def _validate_sheet_title(self, sheet_title: str) -> str:
        """
        Excelシート名のバリデーションを行い、使用できない文字や不適切な先頭・末尾のシングルクォートを削除する。

        無効な文字: / \ ? * [ ] : （全角記号も含む）
        """
        # 使用できない文字を指定の文字（例: "_"）に置換
        sanitized_title = re.sub(r'[\/\\\?\*\[\]\:／＼？＊［］：]', '_', sheet_title)

        # 先頭と末尾のシングルクォートを削除
        sanitized_title = sanitized_title.strip("'")

        # シート名の長さを31文字以内に制限
        if len(sanitized_title) > 31:
            sanitized_title = sanitized_title[:31]

        return sanitized_title

    def write_to_sheet(self, ws: Workbook, data: Any) -> Workbook:
        """データをシートに書き込む。"""
        if isinstance(data, str):
            # 文字列データの場合
            for line in data.replace('<br>', '\n').split('\n'):
                # テーブルのマークダウン記法を検出
                if line.startswith('|') and line.endswith('|'):
                    # 行が '|', '-', ':'のみで構成されている場合はパス
                    if set(line.replace('|', '').replace('-', '').replace(':', '').strip()) == set():
                        continue
                    # 先頭と末尾の'|'を削除し、'|'で分割
                    cells = line.strip('|').split('|')
                    # 各要素をトリムしてセルに追加
                    ws.append([cell.strip() for cell in cells])
                else:
                    # 通常の行として追加
                    ws.append([line])
        elif isinstance(data, (list, dict)):
            # リストまたは辞書データの場合
            if isinstance(data, dict):
                data = [data]

            # 空のデータチェック
            if not data:
                return ws

            # データ型の検証
            if not all(isinstance(item, dict) for item in data):
                ws.append([str(data)])
                return ws

            # ヘッダーの収集、ヘッダー行の追加（入れ子になった辞書のフィールドも含める）
            headers = []
            for row in data:
                for key, value in row.items():
                    if isinstance(value, dict):
                        for dict_key in value.keys():
                            if dict_key not in headers:
                                headers.append(dict_key)
                    elif isinstance(value, list) and value and isinstance(value[0], dict):
                        for dict_key in value[0].keys():
                            if dict_key not in headers:
                                headers.append(dict_key)
                    elif key not in headers:
                        headers.append(key)
            ws.append(headers)

            # データ行の追加
            for row in data:
                # 辞書型配列を含むフィールドを探す
                array_fields = {
                    key: value for key, value in row.items()
                    if isinstance(value, list) and value and isinstance(value[0], dict)
                }

                if array_fields:
                    max_length = max(len(value)
                                     for value in array_fields.values())

                    for i in range(max_length):
                        row_data = []
                        for header in headers:
                            value = ""
                            # 通常のフィールド（リストも含む）
                            if header in row:
                                value = row[header]
                                # リストの場合（辞書型でない場合）はカンマ区切りの文字列に変換
                                if isinstance(value, list) and (not value or not isinstance(value[0], dict)):
                                    value = ", ".join(str(v) for v in value)
                            # 辞書型配列のフィールド
                            else:
                                for array_field in array_fields.values():
                                    if i < len(array_field) and header in array_field[i]:
                                        value = array_field[i][header]
                                        break
                            row_data.append(value)
                        ws.append(row_data)
                else:
                    # 通常のデータ行の処理
                    row_data = []
                    for header in headers:
                        value = row.get(header, "")
                        if isinstance(value, list) and (not value or not isinstance(value[0], dict)):
                            value = ", ".join(str(v) for v in value)
                        row_data.append(value)
                    ws.append(row_data)

        return self._adjust_column_width(ws)

    def safe_cell_conversion(self, value: Any) -> Any:
        """安全なセルに変換する。"""
        if isinstance(value, str) and len(value) > 32767:  # Excelでの最大文字列長
            return value[:32767]
        return value

    def _adjust_column_width(self, ws):
        """
        ワークシートの列幅を内容に合わせて自動調整する。
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            # その列の全セルをチェック
            for cell in column:
                try:
                    # セルの値を文字列として取得
                    cell_value = str(
                        cell.value) if cell.value is not None else ""
                    # 現在の最大長と比較
                    max_length = max(max_length, len(cell_value))
                except Exception as e:
                    logging.warning(
                        f"Error processing cell {cell.coordinate}: {str(e)}")
                    continue

            # 最小幅と最大幅の制限を設定
            adjusted_width = min(max(max_length + 2, 8), 50)  # 最小8文字、最大50文字

            # 列幅を設定
            ws.column_dimensions[column_letter].width = adjusted_width

    # def _write_formatted_data_to_new_sheets(self, wb: Workbook, data: Union[List[Dict[str, Any]], Dict[str, Any]]) -> None:
    #     """
    #     比較元ファイルと比較先ファイルのデータを項目ごと対応させたデータをファイルごとシートに出力する。

    #     dataのスキーマ:
    #     {
    #     "保証品位や基準品位": [
    #         {
    #             "項目名": "Diameter", // 項目名1（条件含む）
    #             "保証品位や基準品位": "6～8", // 保証品位または基準品位1の数値のみ
    #             "保証品位や基準品位の単位: "mm" // 保証品位または基準品位1の単位のみ
    #         },
    #         ...
    #     ],
    #     "分析結果": [
    #         {
    #             "項目名": "Diameter", // 項目名1（条件含む）
    #             "分析結果": "8", // 分析結果の数値のみ
    #             "分析結果の単位: "mm" // 分析結果の単位のみ
    #         },
    #         ...
    #     ],
    #     }
    #     """
    #     try:
    #         logging.info("Starting write_formatted_data_to_sheets")
    #         logging.info(f"Input data: {data}")

    #         normalized_data = self._normalize_data(data)

    #         # 保証品位シートの作成と書き込み
    #         ws_source = wb.create_sheet(title="比較元ファイル(整形済み)")
    #         source_data = normalized_data.get("保証品位や基準品位", [])

    #         if not source_data:
    #             logging.warning("Source data is empty")
    #             return

    #         logging.info(f"Source data: {source_data}")

    #         # ヘッダーの書き込み
    #         headers = list(source_data[0].keys())
    #         for col, header in enumerate(headers, 1):
    #             ws_source.cell(row=1, column=col, value=header)

    #         # データの書き込み
    #         for row, item in enumerate(source_data, 2):
    #             for col, key in enumerate(headers, 1):
    #                 ws_source.cell(row=row, column=col, value=item.get(key, ""))

    #         # 列幅の自動調整
    #         self._adjust_column_width(ws_source)

    #         # 分析結果シートの作成と書き込み
    #         ws_target = wb.create_sheet(title="比較先ファイル(整形済み)")
    #         target_data = normalized_data.get("分析結果", [])

    #         if not target_data:
    #             logging.warning("Target data is empty")
    #             return

    #         logging.info(f"Target data: {target_data}")

    #         # ヘッダーの書き込み
    #         headers = list(target_data[0].keys())
    #         for col, header in enumerate(headers, 1):
    #             ws_target.cell(row=1, column=col, value=header)

    #         # データの書き込み
    #         for row, item in enumerate(target_data, 2):
    #             for col, key in enumerate(headers, 1):
    #                 ws_target.cell(row=row, column=col, value=item.get(key, ""))

    #         # 列幅の自動調整
    #         self._adjust_column_width(ws_target)

    #     except Exception as e:
    #         logging.error(f"Error in write_formatted_data_to_sheets: {traceback.format_exc()}")
    #         raise


###############
# OCR処理クラス #
###############
class DocumentProcessor(BaseDocumentProcessor):
    """
    汎用業務パターンの文書処理クラス。

    業務パターンが「船積書類文字起こし」、「複数ファイル突合」以外の場合に使用する。
    """

    def __init__(self, business_pattern: str = "ocr_results", upn: str = "unknown"):
        self.upn = upn
        self.business_pattern = business_pattern

    async def process(self, ocr_contents, system_prompts):
        """
        OCRで読み取った書類のテキストから、情報を抽出し、Excelファイルに出力する。
        """
        try:
            # AOAIでテキストを分析し、構造化された情報を抽出する
            all_results = await self._analyze_ocr_contents_with_aoai(ocr_contents, system_prompts)

            excel_service = ExcelService()
            excel_files = []

            # 各OCRコンテンツに対して個別のExcelファイルを生成
            for idx, ocr_content in enumerate(ocr_contents):
                # 新しいワークブックを作成
                wb = excel_service.create_workbook()

                # OCR結果シートの作成
                wb, sheet_name = excel_service.create_unique_sheet(
                    wb, ocr_content["fileName"][:31])
                ws = wb[sheet_name]
                excel_service.write_to_sheet(ws, ocr_content["content"])

                # このファイルに関連する処理結果シートの作成
                document_results = all_results["documents"][idx]
                for result_item in document_results["results"]:
                    wb, sheet_name = excel_service.create_unique_sheet(
                        wb, result_item["title"])
                    ws = wb[sheet_name]
                    excel_service.write_to_sheet(ws, result_item["result"])

                # プロンプトシートの作成
                for prompt in all_results["prompts"]:
                    wb, sheet_name = excel_service.create_unique_sheet(
                        wb, f"{prompt['title']}指示文")
                    ws = wb[sheet_name]
                    excel_service.write_to_sheet(ws, prompt["prompt"])

                # Excelファイルを一時的にメモリに保存
                with io.BytesIO() as excel_stream:
                    wb.save(excel_stream)
                    excel_stream.seek(0)
                    excel_data = excel_stream.getvalue()

                # ファイル名の生成
                jst = ZoneInfo('Asia/Tokyo')
                timestamp = datetime.now(jst).strftime("%Y%m%d%H%M%S")
                file_name = f"{self.business_pattern}_{ocr_content['fileName']}_{timestamp}.xlsx"
                blob_name = f"{self.upn}/{file_name}"

                # Azure Blob Storageに保存
                await self.upload_blob(blob_name, excel_data)

                # Base64エンコード
                excel_base64 = base64.b64encode(excel_data).decode('utf-8')

                # 結果をリストに追加
                excel_files.append({
                    "fileName": file_name,
                    "data": excel_base64
                })

            return excel_files

        except Exception as e:
            logging.error(
                f"Error in DocumentProcessor.process: {traceback.format_exc()}")
            raise

    async def _analyze_ocr_contents_with_aoai(self, ocr_contents: List[Dict], system_prompts: List[Dict]) -> Dict[str, Any]:
        """
        OCR結果をAOAIで分析し、構造化された情報を抽出し、プロンプトと共に返す。

        Returns:
            Dict[str, Any]: {
                "documents": [  # 全てのOCR contentsの結果
                    {
                        "filename": "ファイル1.pdf",
                        "results": [
                            {
                                "title": "プロンプト1のタイトル_ファイル1.pdf",
                                "result": "プロンプト1の結果"
                            },
                            {
                                "title": "プロンプト2のタイトル_ファイル1.pdf",
                                "result": "プロンプト2の結果"
                            }
                        ]
                    },
                    ...
                ],
                "prompts": List[Dict] # システムプロンプトのリスト(全てのファイルに対して共通のプロンプトを使用するため"documents"とは別に格納)
            }
        """
        all_results = []
        for ocr_content in ocr_contents:
            # 各プロンプトの処理を並列実行
            tasks = [
                self._process_single_prompt(
                    ocr_content["content"], prompt_config, model_gpt4_1_mini, json_mode=True)
                for prompt_config in system_prompts
            ]
            content_results = await asyncio.gather(*tasks)

            # 結果とタイトルをペアにする
            results_with_titles = [
                {
                    "title": f"{prompt['title']}_{ocr_content['fileName']}",
                    "result": result
                }
                for prompt, result in zip(system_prompts, content_results)
            ]
            all_results.append({
                "filename": ocr_content["fileName"],
                "results": results_with_titles
            })

        return {
            "documents": all_results,
            "prompts": system_prompts
        }


class ShippingDocumentTextExtractor(BaseDocumentProcessor):
    """
    船積書類文字起こし処理クラス。

    OCRで読み取った船積書類のテキストから、
    L/C、B/L、Invoice、P/L、COO、COA、COQの情報を抽出し、Excelファイルに出力する。
    """

    def __init__(self, upn: str = "unknown"):
        self.upn = upn
        self.key_mappings = {
            "PO": ["PO", "P/O", "Purchase Order", "PURCHASE ORDER"],
            "LC": ["L/C", "L／C", "LC"],
            "BL": ["B/L", "B／L", "BL"],
            "PL": ["P/L", "P／L", "PL"],
            "INV": ["Invoice", "Inv", "INVOICE", "INV"],
            "COO": ["COO"],
            "COA": ["COA"],
            "COQ": ["COQ"]
        }

    async def process(self, ocr_contents: List[Dict], system_prompts: List[Dict]) -> Dict:
        """船積書類のOCR結果を判別結果を元にページ分割し、種別ごとに処理する。"""
        try:
            excel_files = []

            for ocr_content in ocr_contents:
                # OCRテキストをページ単位に分割
                document_pages = self._split_content_by_pages(
                    ocr_content["content"])

                # 判別された帳票と該当ページの辞書を作成
                document_page_mappings = []
                for key, value in ocr_content["judgeResult"].items():
                    if key not in ["その他", "ファイル名"] and isinstance(value, list):
                        standard_key = self._map_to_standard_key(
                            key)  # マッピングされた標準キーを取得
                        prompt_config = self._get_prompt(
                            system_prompts, standard_key)
                        document_page_mappings.append({
                            "type": key,
                            "contents_page": value,
                            "prompt": prompt_config
                        })

                prompts = []
                all_results = []
                task_prompt_pairs = []
                for item in document_page_mappings:
                    target_pages = item.get("contents_page")
                    prompt = item.get("prompt")

                    # 対象ページがある場合、ページ内容を結合してLLM処理を準備
                    page_contents = []
                    page_contents.append(document_pages[0])  # ファイル名を追加
                    for page_num in target_pages:
                        page_contents.append(
                            document_pages[page_num])  # 対象ページの内容を追加
                    combined_page_content = "\n".join(page_contents)

                    task = self._process_single_prompt(
                        combined_page_content,
                        prompt,
                        model_gpt4_1_mini,
                        json_mode=True
                    )
                    task_prompt_pairs.append((task, prompt))

                # タスクとプロンプトのペアを維持
                tasks = [pair[0] for pair in task_prompt_pairs]
                prompts = [pair[1] for pair in task_prompt_pairs]

                # 全てのタスクを並列実行
                content_results = await asyncio.gather(*tasks, return_exceptions=True)

                # 結果とタイトルをペアにする
                results_with_titles = [
                    {
                        "title": prompt['title'],
                        "result": result
                    }
                    for prompt, result in zip(prompts, content_results)
                ]
                all_results.append({
                    "filename": ocr_content["fileName"],
                    "results": results_with_titles
                })

                excel_service = ExcelService()

                # 新しいワークブックを作成
                wb = excel_service.create_workbook()

                # OCR結果シートの作成
                wb, sheet_name = excel_service.create_unique_sheet(
                    wb, "全文字起こし")
                ws = wb[sheet_name]
                excel_service.write_to_sheet(ws, ocr_content["content"])

                # このファイルに関連する処理結果シートの作成
                for result_item in all_results:
                    for result in result_item["results"]:
                        wb, sheet_name = excel_service.create_unique_sheet(
                            wb, result["title"])
                        ws = wb[sheet_name]
                        excel_service.write_to_sheet(ws, result["result"])

                # プロンプトシートの作成
                for prompt in prompts:
                    wb, sheet_name = excel_service.create_unique_sheet(
                        wb, f"{prompt['title']}指示文")
                    ws = wb[sheet_name]
                    excel_service.write_to_sheet(ws, prompt["prompt"])

                # Excelファイルを一時的にメモリに保存
                with io.BytesIO() as excel_stream:
                    wb.save(excel_stream)
                    excel_stream.seek(0)
                    excel_data = excel_stream.getvalue()

                # ファイル名の生成
                jst = ZoneInfo('Asia/Tokyo')
                timestamp = datetime.now(jst).strftime("%Y%m%d%H%M%S")
                file_name = f"{ocr_content['fileName']}_{timestamp}.xlsx"
                blob_name = f"{self.upn}/{file_name}"

                # Azure Blob Storageに保存
                await self.upload_blob(blob_name, excel_data)

                # Base64エンコード
                excel_base64 = base64.b64encode(excel_data).decode('utf-8')

                # 結果をリストに追加
                excel_files.append({
                    "fileName": file_name,
                    "data": excel_base64
                })

            return excel_files

        except Exception as e:
            logging.error(
                f"Error in ShippingDocumentClassifier.process: {traceback.format_exc()}")
            raise

    def _map_to_standard_key(self, input_key: str) -> str:
        """入力されたキーを標準化された形式に変換する

        Args:
            input_key (str): 入力キー

        Returns:
            str: 標準化されたキー。マッピングが見つからない場合は元のキーを返す
        """
        for standard_key, variants in self.key_mappings.items():
            if input_key in variants:
                return standard_key
        return input_key


# class DocumentComparisonProcessor(BaseDocumentProcessor):
#     """複数ファイル比較・突合処理クラス"""
#     def __init__(self, upn: str = "unknown"):
#         self.upn = upn
#         self.json_mode = True
#         self.excel_service = ExcelService("複数ファイル突合")

#     async def process(self, ocr_contents: List[Dict], system_prompts: List[Dict]) -> Dict:
#         """
#         OCRコンテンツから文書の比較を実行する。
#         """
#         try:
#             # データのフォーマット処理
#             formatted_data = await self._format_documents(ocr_contents)

#             # 比較可能なデータの準備
#             comparable_data = await self._prepare_comparison(ocr_contents)

#             # 文書の比較実行
#             comparison_result = await self._compare_documents(
#                 comparable_data["result"],
#                 system_prompts[0]["prompt"]
#             )

#             responses = {
#                 'formatted_data': formatted_data[0],
#                 'comparable_data': comparable_data[0],
#                 'comparison_result': comparison_result[0]
#             }

#             # Excelファイル生成
#             excel_data = await self.excel_service.generate(
#                 ocr_contents,
#                 system_prompts,
#                 responses
#             )

#             # ファイル名の生成
#             jst = ZoneInfo('Asia/Tokyo')
#             timestamp = datetime.now(jst).strftime("%Y%m%d%H%M%S")
#             file_name = f"複数ファイル突合_{timestamp}.xlsx"
#             blob_name = f"{self.upn}/{file_name}"

#             # Azure Blob Storageに保存
#             await self.upload_blob(blob_name, excel_data)

#             # Base64エンコード
#             excel_base64 = base64.b64encode(excel_data).decode('utf-8')

#             return file_name, excel_base64

#         except Exception as e:
#             logging.error(f"Error in DocumentComparisonProcessor.process: {traceback.format_exc()}")
#             raise

#     async def _format_documents(self, ocr_contents: List[Dict]) -> List[Dict]:
#         """文書のフォーマット処理"""
#         combined_content = json.dumps({
#             "ファイル１": ocr_contents[0]["content"],
#             "ファイル２": ocr_contents[1]["content"]
#         }, ensure_ascii=False)

#         result = await self.generate_response(
#             content=combined_content,
#             system_prompt=coa_comparison_sub_prompt_list[0]["prompt"],
#             model_name=model_gpt4_1_mini,
#             json_mode=self.json_mode
#         )

#         return [{
#             "title": "formatted_data",
#             "result": result
#         }]

#     async def _prepare_comparison(self, ocr_contents: List[Dict]) -> List[Dict]:
#         """2つのファイルのデータを比較できるように対応させ、1つのデータを生成する。"""
#         combined_content = json.dumps({
#             "ファイル１": ocr_contents[0]["content"],
#             "ファイル２": ocr_contents[1]["content"]
#         }, ensure_ascii=False)

#         result = await self.generate_response(
#             content=combined_content,
#             system_prompt=coa_comparison_sub_prompt_list[1]["prompt"],
#             model_name=model_gpt4_1,
#             json_mode=self.json_mode
#         )

#         return [{
#             "title": "comparable_data",
#             "result": result
#         }]

#     async def _compare_documents(self, comparable_data: Dict, system_prompt: str) -> List[Dict]:
#         """文書を比較実行する"""
#         result = await self.generate_response(
#             content=json.dumps(comparable_data, ensure_ascii=False),
#             system_prompt=system_prompt,
#             model_name=model_gpt4_1,
#             json_mode=self.json_mode
#         )

#         return [{
#             "title": "comparison_result",
#             "result": result
#         }]


class ShippingDocumentTypeIdentifier(BaseDocumentProcessor):
    """船積書類分類処理クラス"""

    def __init__(self):
        self.json_mode = True

    async def process(self, ocr_contents: List[Dict], system_prompts: List[Dict]) -> List[Dict]:
        """
        OCRコンテンツから書類種別を判別する。

        Args:
            ocr_contents: OCR結果のリスト
            system_prompts: 分類用のプロンプトリスト

        Returns:
            List[Dict]: 各書類の分類結果
        """
        try:
            tasks = []
            for ocr_content in ocr_contents:
                for prompt_config in system_prompts:
                    tasks.append(
                        self._process_single_prompt(
                            content=ocr_content["content"],
                            prompt_config=prompt_config,
                            model_name=model_o4_mini,
                            json_mode=self.json_mode
                        )
                    )

            # 全てのタスクを並列実行
            results = await asyncio.gather(*tasks, return_exceptions=True)

            return results

        except Exception as e:
            logging.error(
                f"Error in DocumentClassifier.process: {traceback.format_exc()}")
            raise
